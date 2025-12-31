"""
Автоматический поиск групп и каналов в Telegram
Использует Telethon для работы с Telegram API
"""

import asyncio
from datetime import datetime
from typing import List, Dict, Set

from telethon import TelegramClient
from telethon.tl.types import Channel, Chat
from telethon.tl.functions.contacts import SearchRequest
from telethon.tl.functions.channels import GetFullChannelRequest, JoinChannelRequest, GetParticipantRequest, GetForumTopicsRequest
from telethon.tl.functions.messages import GetFullChatRequest, ImportChatInviteRequest
from telethon.tl.types import ChannelParticipantSelf, Channel, Chat
from telethon.errors import UsernameInvalidError, UsernameNotOccupiedError, InviteHashExpiredError, UserBannedInChannelError, FloodWaitError, UserNotParticipantError

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


class TelegramSearcher:
    def __init__(self, api_id: int, api_hash: str, session_name: str = 'telegram_session', search_delay: float = 1.0):
        """
        Инициализация клиента Telegram
        
        Args:
            api_id: API ID из my.telegram.org
            api_hash: API Hash из my.telegram.org
            session_name: Имя файла сессии (для сохранения авторизации)
            search_delay: Задержка между поисковыми запросами в секундах (0 = без задержки)
        """
        self.api_id = api_id
        self.api_hash = api_hash
        self.session_name = session_name
        self.search_delay = search_delay
        self.client = TelegramClient(session_name, api_id, api_hash)
        # Хранилище для результатов (для сохранения при прерывании)
        self.current_results = {'groups': [], 'channels': []}
    
    @staticmethod
    def transliterate(text: str) -> str:
        """
        Транслитерация русских букв в английские
        
        Args:
            text: Текст для транслитерации
            
        Returns:
            Транслитерированный текст
        """
        translit_map = {
            'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
            'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
            'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
            'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
            'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
            'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo',
            'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
            'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
            'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
            'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
        }
        
        result = []
        for char in text:
            result.append(translit_map.get(char, char))
        return ''.join(result)
    
    @staticmethod
    def generate_search_queries(keywords: List[str], cities: List[str] = None) -> List[str]:
        """
        Генерирует комбинации поисковых запросов:
        - Оригинальные ключевые слова
        - Транслитерированные ключевые слова
        - Ключевые слова + города
        - Транслитерированные ключевые слова + города
        
        Args:
            keywords: Список ключевых слов
            cities: Список городов (опционально)
            
        Returns:
            Список всех комбинаций для поиска
        """
        queries: Set[str] = set()
        
        # Добавляем оригинальные ключевые слова
        for keyword in keywords:
            if keyword.strip():
                queries.add(keyword.strip())
        
        # Добавляем транслитерированные ключевые слова
        for keyword in keywords:
            if keyword.strip():
                translit = TelegramSearcher.transliterate(keyword.strip())
                if translit != keyword.strip():  # Добавляем только если изменилось
                    queries.add(translit)
        
        # Добавляем комбинации с городами
        if cities:
            for keyword in keywords:
                if not keyword.strip():
                    continue
                
                for city in cities:
                    if not city.strip():
                        continue
                    
                    # Оригинальное слово + город
                    queries.add(f"{keyword.strip()} {city.strip()}")
                    queries.add(f"{city.strip()} {keyword.strip()}")
                    
                    # Транслит слова + город
                    translit_keyword = TelegramSearcher.transliterate(keyword.strip())
                    if translit_keyword != keyword.strip():
                        queries.add(f"{translit_keyword} {city.strip()}")
                        queries.add(f"{city.strip()} {translit_keyword}")
                    
                    # Слово + транслит города
                    translit_city = TelegramSearcher.transliterate(city.strip())
                    if translit_city != city.strip():
                        queries.add(f"{keyword.strip()} {translit_city}")
                        queries.add(f"{translit_city} {keyword.strip()}")
                    
                    # Транслит слова + транслит города
                    if translit_keyword != keyword.strip() and translit_city != city.strip():
                        queries.add(f"{translit_keyword} {translit_city}")
                        queries.add(f"{translit_city} {translit_keyword}")
        
        return sorted(list(queries))
        
    async def connect(self):
        """Подключение к Telegram"""
        await self.client.start()
        print("✅ Успешно подключено к Telegram")
    
    async def _get_members_count(self, entity) -> int:
        """
        Получает количество участников группы/канала
        
        Args:
            entity: Объект Channel или Chat
            
        Returns:
            Количество участников или 0 если не удалось получить
        """
        try:
            if isinstance(entity, Channel):
                # Для каналов и супергрупп используем GetFullChannelRequest
                try:
                    full_channel = await self.client(GetFullChannelRequest(entity))
                    if hasattr(full_channel, 'full_chat'):
                        count = getattr(full_channel.full_chat, 'participants_count', None)
                        if count is not None:
                            return count
                except Exception:
                    pass
                
                # Альтернативный способ - через get_entity
                try:
                    full_info = await self.client.get_entity(entity)
                    count = getattr(full_info, 'participants_count', None)
                    if count is not None:
                        return count
                except Exception:
                    pass
                
                # Если ничего не помогло, пробуем получить из самого объекта
                count = getattr(entity, 'participants_count', None)
                if count is not None:
                    return count
                    
            elif isinstance(entity, Chat):
                # Для обычных групп используем GetFullChatRequest
                try:
                    full_chat = await self.client(GetFullChatRequest(entity.id))
                    if hasattr(full_chat, 'full_chat'):
                        count = getattr(full_chat.full_chat, 'participants_count', None)
                        if count is not None:
                            return count
                except Exception:
                    pass
                
                # Альтернативный способ
                count = getattr(entity, 'participants_count', None)
                if count is not None:
                    return count
                    
        except Exception:
            pass
        
        return 0
        
    async def search_channels_and_groups(self, keywords: List[str], limit_per_keyword: int = 50) -> Dict:
        """
        Поиск групп и каналов по ключевым словам
        
        Args:
            keywords: Список ключевых слов для поиска
            limit_per_keyword: Максимальное количество результатов на одно ключевое слово
            
        Returns:
            Словарь с найденными группами и каналами
        """
        all_groups = []
        all_channels = []
        seen_ids = set()  # Для избежания дубликатов
        
        # Обновляем текущие результаты в классе (для возможности сохранения при прерывании)
        self.current_results = {'groups': [], 'channels': []}
        
        print(f"\n🔍 Начинаю поиск по {len(keywords)} ключевым словам...")
        
        for keyword in keywords:
            print(f"\n📝 Ищу: '{keyword}'...")
            try:
                # Метод 1: Поиск через глобальный поиск Telegram
                try:
                    # Используем SearchRequest для поиска контактов/чатов
                    results = await self.client(SearchRequest(
                        q=keyword,
                        limit=limit_per_keyword
                    ))
                    
                    # Обработка результатов
                    for result in results.chats:
                        if not isinstance(result, (Channel, Chat)):
                            continue
                        
                        # Проверяем на дубликаты
                        entity_id = result.id
                        if entity_id in seen_ids:
                            continue
                        seen_ids.add(entity_id)
                        
                        # Получаем количество участников
                        members_count = await self._get_members_count(result)
                        
                        entity_info = {
                            'id': entity_id,
                            'title': result.title,
                            'username': getattr(result, 'username', None),
                            'members_count': members_count,
                            'keyword': keyword
                        }
                        
                        if isinstance(result, Channel):
                            if result.broadcast:
                                all_channels.append(entity_info)
                                self.current_results['channels'].append(entity_info)  # Сохраняем в классе
                                username_str = f" (@{entity_info['username']})" if entity_info['username'] else ""
                                members_str = f" [{members_count:,} подписчиков]" if members_count > 0 else ""
                                print(f"  📢 Канал: {result.title}{username_str}{members_str}")
                            else:
                                all_groups.append(entity_info)
                                self.current_results['groups'].append(entity_info)  # Сохраняем в классе
                                username_str = f" (@{entity_info['username']})" if entity_info['username'] else ""
                                members_str = f" [{members_count:,} участников]" if members_count > 0 else ""
                                print(f"  👥 Группа: {result.title}{username_str}{members_str}")
                        elif isinstance(result, Chat):
                            all_groups.append(entity_info)
                            self.current_results['groups'].append(entity_info)  # Сохраняем в классе
                            members_str = f" [{members_count:,} участников]" if members_count > 0 else ""
                            print(f"  👥 Группа: {result.title}{members_str}")
                            
                except Exception as e:
                    error_msg = str(e)
                    # Проверяем, это flood wait?
                    if "wait of" in error_msg and "seconds" in error_msg:
                        # Извлекаем время ожидания
                        try:
                            wait_seconds = int(error_msg.split("wait of")[1].split("seconds")[0].strip())
                            wait_minutes = wait_seconds / 60
                            wait_hours = wait_minutes / 60
                            print(f"  ⚠️ Flood Wait: Telegram требует подождать {wait_seconds} секунд (~{wait_minutes:.0f} минут, ~{wait_hours:.1f} часов)")
                            print(f"  💡 Рекомендация: Увеличьте SEARCH_DELAY в config.py до 2.0-3.0 секунд")
                            print(f"  💡 Или уменьшите количество городов/ключевых слов")
                            if wait_seconds > 3600:  # Больше часа
                                print(f"  ⏸️ Пропускаю этот запрос и продолжаю с другими...")
                                continue
                        except:
                            pass
                    else:
                        print(f"  ⚠️ Ошибка при поиске через SearchRequest: {e}")
                    
                    # Метод 2: Поиск по уже известным диалогам
                    print(f"  🔄 Пробую альтернативный метод поиска...")
                    try:
                        async for dialog in self.client.iter_dialogs(limit=200):
                            if not isinstance(dialog.entity, (Channel, Chat)):
                                continue
                            
                            title = dialog.entity.title.lower()
                            if keyword.lower() not in title:
                                continue
                            
                            entity_id = dialog.entity.id
                            if entity_id in seen_ids:
                                continue
                            seen_ids.add(entity_id)
                            
                            # Получаем количество участников
                            members_count = await self._get_members_count(dialog.entity)
                            
                            entity_info = {
                                'id': entity_id,
                                'title': dialog.entity.title,
                                'username': getattr(dialog.entity, 'username', None),
                                'members_count': members_count,
                                'keyword': keyword
                            }
                            
                            if isinstance(dialog.entity, Channel):
                                if dialog.entity.broadcast:
                                    all_channels.append(entity_info)
                                    self.current_results['channels'].append(entity_info)  # Сохраняем в классе
                                    username_str = f" (@{entity_info['username']})" if entity_info['username'] else ""
                                    members_str = f" [{members_count:,} подписчиков]" if members_count > 0 else ""
                                    print(f"  📢 Канал: {dialog.entity.title}{username_str}{members_str}")
                                else:
                                    all_groups.append(entity_info)
                                    self.current_results['groups'].append(entity_info)  # Сохраняем в классе
                                    username_str = f" (@{entity_info['username']})" if entity_info['username'] else ""
                                    members_str = f" [{members_count:,} участников]" if members_count > 0 else ""
                                    print(f"  👥 Группа: {dialog.entity.title}{username_str}{members_str}")
                            elif isinstance(dialog.entity, Chat):
                                all_groups.append(entity_info)
                                self.current_results['groups'].append(entity_info)  # Сохраняем в классе
                                members_str = f" [{members_count:,} участников]" if members_count > 0 else ""
                                print(f"  👥 Группа: {dialog.entity.title}{members_str}")
                    except Exception as e2:
                        print(f"  ❌ Ошибка при альтернативном поиске: {e2}")
                    
            except Exception as e:
                print(f"  ❌ Ошибка при поиске '{keyword}': {e}")
            
            # Небольшая задержка между запросами (чтобы не получить flood wait)
            # Можно уменьшить до 0.5-1 секунды, но есть риск получить flood wait
            delay = getattr(self, 'search_delay', 1.0)
            if delay > 0:
                await asyncio.sleep(delay)
        
        print(f"\n✅ Поиск завершен!")
        print(f"   Найдено групп: {len(all_groups)}")
        print(f"   Найдено каналов: {len(all_channels)}")
        
        # Обновляем финальные результаты в классе
        self.current_results = {
            'groups': all_groups,
            'channels': all_channels
        }
        
        return {
            'groups': all_groups,
            'channels': all_channels
        }
    
    def save_to_excel(self, groups: List[Dict], channels: List[Dict], 
                     groups_file: str = 'telegram_groups.xlsx',
                     channels_file: str = 'telegram_channels.xlsx'):
        """
        Сохранение результатов в Excel файлы
        
        Args:
            groups: Список найденных групп
            channels: Список найденных каналов
            groups_file: Имя файла для групп
            channels_file: Имя файла для каналов
        """
        # Сохранение групп
        if groups:
            wb_groups = Workbook()
            ws_groups = wb_groups.active
            ws_groups.title = "Telegram Groups"
            
            # Заголовки
            headers = ['ID', 'Название', 'Username', 'Количество участников', 'Ключевое слово']
            ws_groups.append(headers)
            
            # Стилизация заголовков
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws_groups[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Данные
            for group in groups:
                members = group['members_count']
                # Форматируем количество участников
                if isinstance(members, int) and members > 0:
                    members_str = f"{members:,}".replace(',', ' ')  # Разделитель тысяч
                else:
                    members_str = str(members) if members else 'N/A'
                
                ws_groups.append([
                    group['id'],
                    group['title'],
                    group['username'] or 'N/A',
                    members_str,
                    group['keyword']
                ])
            
            # Автоматическая ширина колонок
            for column in ws_groups.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_groups.column_dimensions[column_letter].width = adjusted_width
            
            wb_groups.save(groups_file)
            print(f"✅ Группы сохранены в: {groups_file}")
        else:
            print("⚠️ Группы не найдены")
        
        # Сохранение каналов
        if channels:
            wb_channels = Workbook()
            ws_channels = wb_channels.active
            ws_channels.title = "Telegram Channels"
            
            # Заголовки
            headers = ['ID', 'Название', 'Username', 'Количество подписчиков', 'Ключевое слово']
            ws_channels.append(headers)
            
            # Стилизация заголовков
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws_channels[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Данные
            for channel in channels:
                members = channel['members_count']
                # Форматируем количество подписчиков
                if isinstance(members, int) and members > 0:
                    members_str = f"{members:,}".replace(',', ' ')  # Разделитель тысяч
                else:
                    members_str = str(members) if members else 'N/A'
                
                ws_channels.append([
                    channel['id'],
                    channel['title'],
                    channel['username'] or 'N/A',
                    members_str,
                    channel['keyword']
                ])
            
            # Автоматическая ширина колонок
            for column in ws_channels.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_channels.column_dimensions[column_letter].width = adjusted_width
            
            wb_channels.save(channels_file)
            print(f"✅ Каналы сохранены в: {channels_file}")
        else:
            print("⚠️ Каналы не найдены")
    
    async def disconnect(self):
        """Отключение от Telegram"""
        await self.client.disconnect()
        print("👋 Отключено от Telegram")
    
    @staticmethod
    def read_groups_from_excel(filename: str) -> List[Dict]:
        """
        Чтение групп из Excel файла (поддерживает разные форматы: обычные группы и ready_groups)
        
        Args:
            filename: Путь к Excel файлу
            
        Returns:
            Список словарей с информацией о группах
        """
        groups = []
        try:
            wb = load_workbook(filename)
            ws = wb.active
            
            # Определяем формат файла по заголовкам
            headers = [cell.value for cell in ws[1]]
            is_ready_format = 'Статус' in headers or 'Сообщение' in headers
            
            # Пропускаем заголовок (первая строка)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Пропускаем только полностью пустые строки (нет ни ID, ни username)
                if not row[0] and not row[2]:  # Нет ID и нет username
                    continue
                
                if is_ready_format:
                    # Формат ready_groups: ID, Название, Username, Участников, Статус, Сообщение, Действие, Ключевое слово, Родительская группа
                    group_info = {
                        'id': int(row[0]) if row[0] and str(row[0]) != 'N/A' else None,
                        'title': str(row[1]) if row[1] else 'N/A',
                        'username': str(row[2]) if row[2] and str(row[2]) != 'N/A' else None,
                        'members_count': row[3] if len(row) > 3 and row[3] != 'N/A' else None,
                        'keyword': row[7] if len(row) > 7 else (row[4] if len(row) > 4 else None),
                        'check_status': 'ready',  # Все группы в ready_groups готовы
                        'check_message': str(row[5]) if len(row) > 5 else 'Готово к рассылке',
                        'check_action': str(row[6]) if len(row) > 6 else 'none',
                        'parent_group': str(row[8]) if len(row) > 8 and row[8] != 'N/A' else None
                    }
                else:
                    # Обычный формат: ID, Название, Username, Количество участников, Ключевое слово
                    # Обработка для файлов, созданных из текста (ID может быть пустым)
                    group_id = None
                    if row[0]:
                        try:
                            if isinstance(row[0], (int, float)):
                                group_id = int(row[0])
                            elif str(row[0]).strip() and str(row[0]) != 'N/A':
                                group_id = int(row[0])
                        except (ValueError, TypeError):
                            pass
                    
                    username = None
                    if row[2] and str(row[2]).strip() and str(row[2]) != 'N/A':
                        username = str(row[2]).strip()
                    
                    group_info = {
                        'id': group_id,
                        'title': str(row[1]) if row[1] else (username or 'N/A'),
                        'username': username,
                        'members_count': row[3] if len(row) > 3 and row[3] else None,
                        'keyword': row[4] if len(row) > 4 else None
                    }
                
                groups.append(group_info)
        except Exception as e:
            print(f"❌ Ошибка чтения файла {filename}: {e}")
        
        return groups
    
    async def check_group_access(self, group_info: Dict, stop_event=None) -> Dict:
        """
        Проверка доступа к группе и возможности отправки сообщений
        
        Args:
            group_info: Словарь с информацией о группе (id, title, username)
            stop_event: threading.Event для остановки проверки
            
        Returns:
            Словарь с результатами проверки:
            {
                'status': 'ready'/'pending'/'unavailable'/'error',
                'message': 'Описание статуса',
                'action_taken': 'joined'/'request_sent'/'none'
            }
        """
        if stop_event and stop_event.is_set():
            return {'status': 'stopped', 'message': 'Проверка остановлена'}
        
        group_id = group_info.get('id')
        username = group_info.get('username')
        title = group_info.get('title', 'Unknown')
        
        try:
            # Пытаемся получить информацию о группе
            entity = None
            
            if username:
                try:
                    entity = await self.client.get_entity(username)
                except (UsernameInvalidError, UsernameNotOccupiedError):
                    # Username недействителен или не существует
                    return {
                        'status': 'unavailable',
                        'message': f'Группа недоступна (неверный username)',
                        'action_taken': 'none'
                    }
            elif group_id:
                try:
                    entity = await self.client.get_entity(group_id)
                except Exception as e:
                    return {
                        'status': 'unavailable',
                        'message': f'Группа недоступна: {str(e)}',
                        'action_taken': 'none'
                    }
            else:
                return {
                    'status': 'unavailable',
                    'message': 'Нет ID или username для группы',
                    'action_taken': 'none'
                }
            
            if not entity:
                return {
                    'status': 'unavailable',
                    'message': 'Не удалось получить информацию о группе',
                    'action_taken': 'none'
                }
            
            # Инициализируем action_taken
            action_taken = 'none'
            
            # ВСЕГДА проверяем участника строгим методом
            is_member = await self._check_membership_strict(entity, title)
            
            # Если не участник - ВСЕГДА пытаемся вступить
            if not is_member:
                print(f"🔍 [{title}] Пользователь НЕ является участником, пытаюсь вступить...")
                action_taken = await self._join_group(entity, username, title)
                print(f"📝 [{title}] Результат вступления: {action_taken}")
                
                if action_taken == 'joined':
                    # Проверяем еще раз после вступления строгим методом
                    await asyncio.sleep(3)  # Увеличиваем задержку для синхронизации
                    is_member = await self._check_membership_strict(entity, title)
                    print(f"✅ [{title}] Проверка после вступления: is_member={is_member}")
                
                # Если все еще не участник после попытки вступления
                if not is_member:
                    return {
                        'status': 'pending',
                        'message': f'Требуется вступление в группу (запрос отправлен или требуется одобрение)',
                        'action_taken': action_taken
                    }
            else:
                print(f"✅ [{title}] Пользователь УЖЕ является участником группы")
            
            # Проверяем возможность отправки сообщений
            if is_member:
                try:
                    # Пытаемся получить права на отправку сообщений
                    if isinstance(entity, Channel):
                        full_info = await self.client(GetFullChannelRequest(entity))
                        # Проверяем, можем ли отправлять сообщения
                        can_send = not getattr(full_info.full_chat, 'default_banned_rights', None) or \
                                  not getattr(full_info.full_chat.default_banned_rights, 'send_messages', False)
                    else:
                        full_info = await self.client(GetFullChatRequest(entity.id))
                        can_send = True  # Для обычных чатов обычно можно
                    
                    if can_send:
                        return {
                            'status': 'ready',
                            'message': 'Готово к рассылке',
                            'action_taken': action_taken
                        }
                    else:
                        return {
                            'status': 'pending',
                            'message': 'В группе, но нет прав на отправку сообщений',
                            'action_taken': action_taken
                        }
                except Exception as e:
                    # Если не можем проверить права, но мы в группе, считаем готовой
                    return {
                        'status': 'ready',
                        'message': 'В группе (права не проверены)',
                        'action_taken': action_taken
                    }
            
        except FloodWaitError as e:
            wait_time = e.seconds
            return {
                'status': 'error',
                'message': f'Flood wait: нужно подождать {wait_time} секунд',
                'action_taken': 'none'
            }
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Ошибка: {str(e)}',
                'action_taken': 'none'
            }
        
        return {
            'status': 'unavailable',
            'message': 'Не удалось определить статус',
            'action_taken': 'none'
        }
    
    async def _check_membership_strict(self, entity, title="") -> bool:
        """
        СТРОГАЯ проверка, является ли пользователь участником группы/канала
        Использует несколько методов для надежности
        
        Returns:
            True если участник, False если нет
        """
        try:
            me = await self.client.get_me()
            
            # Метод 1: Для каналов - проверяем через GetParticipantRequest (самый надежный)
            if isinstance(entity, Channel):
                try:
                    participant = await self.client(GetParticipantRequest(entity, me))
                    # Если получили информацию о себе как участнике - мы участники
                    if isinstance(participant.participant, ChannelParticipantSelf):
                        print(f"✅ [{title}] Проверка через GetParticipantRequest: УЧАСТНИК")
                        return True
                    else:
                        print(f"❌ [{title}] Проверка через GetParticipantRequest: НЕ участник")
                        return False
                except UserNotParticipantError:
                    print(f"❌ [{title}] UserNotParticipantError: НЕ участник")
                    return False
                except Exception as e:
                    error_msg = str(e).lower()
                    if 'not a member' in error_msg or 'not participant' in error_msg or 'user not found' in error_msg:
                        print(f"❌ [{title}] Ошибка указывает на отсутствие: НЕ участник")
                        return False
                    # Если другая ошибка, пробуем другие методы
                    print(f"⚠️ [{title}] GetParticipantRequest ошибка: {e}, пробую другие методы...")
            
            # Метод 2: Проверяем через iter_participants (ищем себя в списке)
            try:
                found_self = False
                async for user in self.client.iter_participants(entity, limit=200):
                    if user.id == me.id:
                        found_self = True
                        break
                if found_self:
                    print(f"✅ [{title}] Найден в списке участников через iter_participants: УЧАСТНИК")
                    return True
                else:
                    print(f"❌ [{title}] НЕ найден в списке участников через iter_participants: НЕ участник")
                    return False
            except Exception as e:
                error_msg = str(e).lower()
                if 'not a member' in error_msg or 'not participant' in error_msg:
                    print(f"❌ [{title}] iter_participants ошибка: НЕ участник")
                    return False
                print(f"⚠️ [{title}] iter_participants ошибка: {e}")
            
            # Метод 3: Для каналов - проверяем через GetFullChannelRequest (может не работать для не-участников)
            if isinstance(entity, Channel):
                try:
                    full_info = await self.client(GetFullChannelRequest(entity))
                    # Если получили без ошибок, но не уверены - проверяем через диалоги
                    # Проверяем, есть ли эта группа в наших диалогах
                    async for dialog in self.client.iter_dialogs():
                        if dialog.entity.id == entity.id:
                            print(f"✅ [{title}] Найдена в диалогах: УЧАСТНИК")
                            return True
                    print(f"❌ [{title}] НЕ найдена в диалогах: НЕ участник")
                    return False
                except Exception as e:
                    error_msg = str(e).lower()
                    if 'not a member' in error_msg or 'not participant' in error_msg:
                        print(f"❌ [{title}] GetFullChannelRequest ошибка: НЕ участник")
                        return False
            
            # Метод 4: Для обычных чатов
            else:
                try:
                    full_info = await self.client(GetFullChatRequest(entity.id))
                    # Проверяем через диалоги
                    async for dialog in self.client.iter_dialogs():
                        if dialog.entity.id == entity.id:
                            print(f"✅ [{title}] Найдена в диалогах: УЧАСТНИК")
                            return True
                    print(f"❌ [{title}] НЕ найдена в диалогах: НЕ участник")
                    return False
                except Exception as e:
                    error_msg = str(e).lower()
                    if 'not a member' in error_msg or 'not participant' in error_msg:
                        print(f"❌ [{title}] GetFullChatRequest ошибка: НЕ участник")
                        return False
            
            # Если все методы не сработали - считаем, что НЕ участник (безопаснее)
            print(f"⚠️ [{title}] Все методы проверки не дали результата, считаю НЕ участником")
            return False
            
        except Exception as e:
            # В случае ошибки считаем, что НЕ участник (безопаснее)
            print(f"⚠️ [{title}] Критическая ошибка при проверке участника: {e}, считаю НЕ участником")
            return False
    
    async def _join_group(self, entity, username=None, title="") -> str:
        """
        Попытка вступить в группу (нажимает кнопку "Присоединиться" или "Подать заявку")
        
        Returns:
            'joined' - успешно вступили
            'request_sent' - отправлен запрос на вступление (кнопка "Подать заявку" обработана)
            'none' - не удалось
        """
        try:
            if isinstance(entity, Channel):
                # Для каналов и супергрупп
                try:
                    print(f"🔄 [{title}] Отправляю JoinChannelRequest (нажимаю кнопку 'Присоединиться'/'Подать заявку')...")
                    await self.client(JoinChannelRequest(entity))
                    await asyncio.sleep(max(self.search_delay, 5.0))  # Минимум 5 секунд между вступлениями
                    
                    # Проверяем строгим методом, вступили ли мы
                    is_member = await self._check_membership_strict(entity, title)
                    if is_member:
                        print(f"✅ [{title}] Успешно вступил в группу (кнопка 'Присоединиться' сработала)")
                        return 'joined'
                    else:
                        # Если не вступили, но запрос прошел - значит отправлен запрос на одобрение
                        # Это означает, что кнопка "Подать заявку" была обработана
                        print(f"⏳ [{title}] Запрос на вступление отправлен (кнопка 'Подать заявку' обработана, ожидается одобрение)")
                        return 'request_sent'
                except FloodWaitError as e:
                    wait_time = e.seconds
                    wait_minutes = wait_time / 60
                    print(f"⏸️ [{title}] ⚠️ Flood Wait: Telegram требует подождать {wait_time} секунд (~{wait_minutes:.1f} минут)")
                    
                    if wait_time > 300:  # Больше 5 минут
                        print(f"⏸️ [{title}] Слишком долгое ожидание ({wait_time} сек), пропускаю эту группу")
                        return 'none'
                    else:
                        print(f"⏳ [{title}] Жду {wait_time} секунд перед следующей попыткой...")
                        await asyncio.sleep(wait_time)
                        # Пробуем еще раз после ожидания
                        try:
                            await self.client(JoinChannelRequest(entity))
                            await asyncio.sleep(3)
                            is_member = await self._check_membership_strict(entity, title)
                            if is_member:
                                return 'joined'
                            else:
                                return 'request_sent'
                        except:
                            return 'request_sent'
                except UserBannedInChannelError:
                    print(f"❌ [{title}] Забанен в канале")
                    return 'none'
                except Exception as e:
                    error_msg = str(e).lower()
                    print(f"⚠️ [{title}] Ошибка при вступлении: {e}")
                    
                    # Проверяем flood wait в тексте ошибки
                    if "wait of" in error_msg and "seconds" in error_msg:
                        try:
                            wait_seconds = int(error_msg.split("wait of")[1].split("seconds")[0].strip())
                            wait_minutes = wait_seconds / 60
                            print(f"⏸️ [{title}] ⚠️ Flood Wait в тексте ошибки: {wait_seconds} секунд (~{wait_minutes:.1f} минут)")
                            
                            if wait_seconds > 300:  # Больше 5 минут
                                print(f"⏸️ [{title}] Слишком долгое ожидание, пропускаю")
                                return 'none'
                            else:
                                print(f"⏳ [{title}] Жду {wait_seconds} секунд...")
                                await asyncio.sleep(wait_seconds)
                                return 'request_sent'  # После ожидания считаем, что запрос отправлен
                        except:
                            pass
                    
                    # Проверяем, это ошибка о необходимости одобрения?
                    if any(keyword in error_msg for keyword in ['request', 'approval', 'invite', 'pending', 'moderation']):
                        print(f"⏳ [{title}] Требуется одобрение администратора (кнопка 'Подать заявку' обработана)")
                        return 'request_sent'
                    
                    # Если ошибка "CHANNEL_PRIVATE" - группа приватная, нужен invite
                    if 'private' in error_msg or 'invite' in error_msg:
                        print(f"⏳ [{title}] Приватная группа, требуется invite-ссылка")
                        return 'request_sent'
                    
                    # Другие ошибки
                    return 'none'
            else:
                # Обычный чат - обычно нельзя вступить автоматически
                print(f"⚠️ [{title}] Обычный чат - автоматическое вступление невозможно")
                return 'none'
        except Exception as e:
            error_msg = str(e).lower()
            print(f"⚠️ [{title}] Исключение при вступлении: {e}")
            
            # Проверяем, это ошибка о необходимости одобрения?
            if any(keyword in error_msg for keyword in ['request', 'approval', 'pending', 'wait', 'moderation']):
                print(f"⏳ [{title}] Требуется одобрение (кнопка 'Подать заявку' обработана)")
                return 'request_sent'
            
            return 'none'
    
    async def process_pending_groups(self, pending_groups: List[Dict], stop_event=None, progress_callback=None) -> Dict:
        """
        Обработка pending групп - автоматическое вступление (нажатие кнопки "Присоединиться")
        
        Args:
            pending_groups: Список групп из pending файла
            stop_event: threading.Event для остановки обработки
            progress_callback: Функция для обновления прогресса (session_id, current, total, message, current_group)
            
        Returns:
            Словарь с результатами:
            {
                'ready_groups': [...],  # Группы, в которые удалось вступить
                'still_pending': [...],  # Группы, которые все еще pending
                'errors': [...]  # Группы с ошибками
            }
        """
        ready_groups = []
        still_pending = []
        errors = []
        
        print(f"\n🔄 Начинаю обработку {len(pending_groups)} pending групп...")
        
        for i, group in enumerate(pending_groups):
            if stop_event and stop_event.is_set():
                print("⏹ Обработка остановлена пользователем")
                break
            
            group_id = group.get('id')
            username = group.get('username')
            title = group.get('title', f"ID: {group_id}")
            
            print(f"\n[{i+1}/{len(pending_groups)}] Обрабатываю: {title}")
            
            # Обновляем прогресс
            if progress_callback:
                progress_callback(i + 1, len(pending_groups), f'Обрабатываю: {title}', title)
            
            try:
                # Получаем entity
                entity = None
                if username:
                    try:
                        entity = await self.client.get_entity(username)
                    except (UsernameInvalidError, UsernameNotOccupiedError):
                        errors.append({
                            **group,
                            'check_status': 'error',
                            'check_message': 'Неверный username',
                            'check_action': 'none'
                        })
                        continue
                elif group_id:
                    try:
                        entity = await self.client.get_entity(group_id)
                    except Exception as e:
                        errors.append({
                            **group,
                            'check_status': 'error',
                            'check_message': f'Не удалось получить entity: {str(e)}',
                            'check_action': 'none'
                        })
                        continue
                else:
                    errors.append({
                        **group,
                        'check_status': 'error',
                        'check_message': 'Нет ID или username',
                        'check_action': 'none'
                    })
                    continue
                
                if not entity:
                    errors.append({
                        **group,
                        'check_status': 'error',
                        'check_message': 'Не удалось получить entity',
                        'check_action': 'none'
                    })
                    continue
                
                # Проверяем, является ли это форумом (группа с темами)
                is_forum = False
                forum_topics = []
                
                if isinstance(entity, Channel):
                    try:
                        full_info = await self.client(GetFullChannelRequest(entity))
                        is_forum = getattr(full_info.full_chat, 'forum', False)
                        
                        if is_forum:
                            print(f"📚 [{title}] Обнаружен форум с темами, получаю список тем...")
                            try:
                                # Получаем темы форума
                                topics_result = await self.client(GetForumTopicsRequest(
                                    channel=entity,
                                    offset_date=0,
                                    offset_id=0,
                                    offset_topic=0,
                                    limit=100
                                ))
                                
                                if hasattr(topics_result, 'topics') and topics_result.topics:
                                    for topic in topics_result.topics:
                                        forum_topics.append({
                                            'id': topic.id,
                                            'title': topic.title,
                                            'parent_group': title,
                                            'parent_group_id': group_id,
                                            'parent_username': username
                                        })
                                    print(f"📚 [{title}] Найдено {len(forum_topics)} тем в форуме")
                            except Exception as e:
                                print(f"⚠️ [{title}] Не удалось получить темы форума: {e}")
                    except Exception as e:
                        print(f"⚠️ [{title}] Ошибка при проверке форума: {e}")
                
                # Обрабатываем основную группу
                is_member = await self._check_membership_strict(entity, title)
                
                if is_member:
                    # Уже участник - переносим в ready
                    print(f"✅ [{title}] Уже участник, переношу в ready")
                    ready_groups.append({
                        **group,
                        'check_status': 'ready',
                        'check_message': 'Уже был участником',
                        'check_action': 'none'
                    })
                    
                    # Если это форум и мы участники - обрабатываем темы
                    if is_forum and forum_topics:
                        print(f"📚 [{title}] Обрабатываю {len(forum_topics)} тем форума...")
                        for topic in forum_topics:
                            topic_title = f"{title} > {topic['title']}"
                            print(f"  📝 Обрабатываю тему: {topic_title}")
                            
                            # Для тем форума проверяем доступ через основную группу
                            # Если мы в основной группе, то имеем доступ к темам
                            ready_groups.append({
                                'id': topic['id'],
                                'title': topic_title,
                                'username': username,  # Username основной группы
                                'members_count': group.get('members_count', 'N/A'),
                                'keyword': group.get('keyword', ''),
                                'check_status': 'ready',
                                'check_message': f'Доступ через форум "{title}"',
                                'check_action': 'forum_topic',
                                'parent_group': title,
                                'parent_group_id': group_id
                            })
                else:
                    # Не участник - пытаемся вступить (нажимаем кнопку "Присоединиться" или "Подать заявку")
                    print(f"🔄 [{title}] Пытаюсь вступить (нажимаю кнопку 'Присоединиться'/'Подать заявку')...")
                    action_taken = await self._join_group(entity, username, title)
                    
                    if action_taken == 'joined':
                        # Проверяем еще раз после вступления
                        await asyncio.sleep(3)
                        is_member = await self._check_membership_strict(entity, title)
                        
                        if is_member:
                            print(f"✅ [{title}] Успешно вступил, переношу в ready")
                            ready_groups.append({
                                **group,
                                'check_status': 'ready',
                                'check_message': 'Успешно вступил',
                                'check_action': 'joined'
                            })
                            
                            # Если это форум и мы вступили - обрабатываем темы
                            if is_forum and forum_topics:
                                print(f"📚 [{title}] Обрабатываю {len(forum_topics)} тем форума...")
                                for topic in forum_topics:
                                    topic_title = f"{title} > {topic['title']}"
                                    ready_groups.append({
                                        'id': topic['id'],
                                        'title': topic_title,
                                        'username': username,
                                        'members_count': group.get('members_count', 'N/A'),
                                        'keyword': group.get('keyword', ''),
                                        'check_status': 'ready',
                                        'check_message': f'Доступ через форум "{title}"',
                                        'check_action': 'forum_topic',
                                        'parent_group': title,
                                        'parent_group_id': group_id
                                    })
                        else:
                            print(f"⏳ [{title}] Вступление не подтверждено, оставляю в pending")
                            still_pending.append({
                                **group,
                                'check_status': 'pending',
                                'check_message': 'Вступление не подтверждено',
                                'check_action': action_taken
                            })
                    elif action_taken == 'request_sent':
                        print(f"⏳ [{title}] Запрос на вступление отправлен (кнопка 'Подать заявку' обработана), оставляю в pending")
                        still_pending.append({
                            **group,
                            'check_status': 'pending',
                            'check_message': 'Запрос на вступление отправлен (ожидается одобрение)',
                            'check_action': 'request_sent'
                        })
                    else:
                        print(f"❌ [{title}] Не удалось вступить, оставляю в pending")
                        still_pending.append({
                            **group,
                            'check_status': 'pending',
                            'check_message': 'Не удалось вступить автоматически',
                            'check_action': action_taken
                        })
                
                # Задержка между группами (увеличена для избежания flood wait)
                if i < len(pending_groups) - 1:
                    await asyncio.sleep(10.0)  # Увеличено до 10 секунд между группами
                    
            except Exception as e:
                print(f"❌ [{title}] Ошибка при обработке: {e}")
                errors.append({
                    **group,
                    'check_status': 'error',
                    'check_message': f'Ошибка: {str(e)}',
                    'check_action': 'none'
                })
        
        print(f"\n✅ Обработка завершена:")
        print(f"   ✅ Готовых: {len(ready_groups)}")
        print(f"   ⏳ Все еще pending: {len(still_pending)}")
        print(f"   ❌ Ошибок: {len(errors)}")
        
        return {
            'ready_groups': ready_groups,
            'still_pending': still_pending + errors,  # Ошибки тоже в pending
            'errors': errors
        }
    
    def save_check_results(self, checked_groups: List[Dict], ready_file: str, pending_file: str):
        """
        Сохранение результатов проверки групп в два Excel файла:
        - ready_file: группы, готовые к рассылке (аккаунт уже в группе)
        - pending_file: группы, требующие действий (запрос отправлен или требуется одобрение)
        
        Args:
            checked_groups: Список словарей с результатами проверки
            ready_file: Имя файла для готовых групп
            pending_file: Имя файла для групп в процессе
        """
        # Разделяем группы на готовые и требующие действий
        ready_groups = []
        pending_groups = []
        other_groups = []
        
        for group in checked_groups:
            status = group.get('check_status', 'unknown')
            if status == 'ready':
                ready_groups.append(group)
            elif status == 'pending':
                pending_groups.append(group)
            else:
                other_groups.append(group)
        
        # Сохраняем готовые группы
        if ready_groups:
            wb_ready = Workbook()
            ws_ready = wb_ready.active
            ws_ready.title = "Ready Groups"
            
            headers = ['ID', 'Название', 'Username', 'Участников', 'Статус', 'Сообщение', 'Действие', 'Ключевое слово', 'Родительская группа']
            ws_ready.append(headers)
            
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws_ready[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            for group in ready_groups:
                ws_ready.append([
                    group.get('id', 'N/A'),
                    group.get('title', 'N/A'),
                    group.get('username') or 'N/A',
                    group.get('members_count', 'N/A'),
                    '✅ Готово к рассылке',
                    group.get('check_message', ''),
                    group.get('check_action', ''),
                    group.get('keyword', ''),
                    group.get('parent_group', 'N/A')  # Для тем форумов
                ])
            
            # Автоматическая ширина колонок
            for column in ws_ready.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_ready.column_dimensions[column_letter].width = adjusted_width
            
            wb_ready.save(ready_file)
            print(f"✅ Готовые группы сохранены в: {ready_file} ({len(ready_groups)} групп)")
        
        # Сохраняем группы в процессе
        if pending_groups or other_groups:
            wb_pending = Workbook()
            ws_pending = wb_pending.active
            ws_pending.title = "Pending Groups"
            
            headers = ['ID', 'Название', 'Username', 'Участников', 'Статус', 'Сообщение', 'Действие', 'Ключевое слово', 'Родительская группа']
            ws_pending.append(headers)
            
            header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws_pending[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Добавляем группы в процессе
            for group in pending_groups:
                status_text = '⏳ Требует действий'
                ws_pending.append([
                    group.get('id', 'N/A'),
                    group.get('title', 'N/A'),
                    group.get('username') or 'N/A',
                    group.get('members_count', 'N/A'),
                    status_text,
                    group.get('check_message', ''),
                    group.get('check_action', ''),
                    group.get('keyword', ''),
                    group.get('parent_group', 'N/A')  # Для тем форумов
                ])
            
            # Добавляем другие группы (недоступные, ошибки)
            for group in other_groups:
                status = group.get('check_status', 'unknown')
                status_text = {
                    'unavailable': '❌ Недоступно',
                    'error': '⚠️ Ошибка',
                    'stopped': '⏹ Остановлено'
                }.get(status, status)
                
                ws_pending.append([
                    group.get('id', 'N/A'),
                    group.get('title', 'N/A'),
                    group.get('username') or 'N/A',
                    group.get('members_count', 'N/A'),
                    status_text,
                    group.get('check_message', ''),
                    group.get('check_action', ''),
                    group.get('keyword', ''),
                    group.get('parent_group', 'N/A')  # Для тем форумов
                ])
            
            # Автоматическая ширина колонок
            for column in ws_pending.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_pending.column_dimensions[column_letter].width = adjusted_width
            
            wb_pending.save(pending_file)
            print(f"✅ Группы в процессе сохранены в: {pending_file} ({len(pending_groups) + len(other_groups)} групп)")
        
        return len(ready_groups), len(pending_groups) + len(other_groups)
    
    async def send_message_to_group(self, group_id, username, title, message_text="", photo_path=None, video_path=None) -> Dict:
        """
        Отправка сообщения в группу
        
        Args:
            group_id: ID группы
            username: Username группы (опционально)
            title: Название группы (для логов)
            message_text: Текст сообщения
            photo_path: Путь к фото (опционально)
            video_path: Путь к видео (опционально)
            
        Returns:
            Словарь с результатом:
            {
                'success': bool,
                'message': str,
                'blocked': bool  # True если заблокирован
            }
        """
        import os
        try:
            # Получаем entity
            entity = None
            if username:
                # Убираем @ если есть, Telethon сам добавит
                clean_username = username.lstrip('@')
                try:
                    entity = await self.client.get_entity(clean_username)
                except (UsernameInvalidError, UsernameNotOccupiedError):
                    return {
                        'success': False,
                        'message': 'Неверный username',
                        'blocked': False
                    }
            elif group_id:
                try:
                    entity = await self.client.get_entity(group_id)
                except Exception as e:
                    return {
                        'success': False,
                        'message': f'Не удалось получить entity: {str(e)}',
                        'blocked': False
                    }
            else:
                return {
                    'success': False,
                    'message': 'Нет ID или username',
                    'blocked': False
                }
            
            # Проверяем, что мы участники
            is_member = await self._check_membership_strict(entity, title)
            if not is_member:
                return {
                    'success': False,
                    'message': 'Не являемся участником группы',
                    'blocked': False
                }
            
            # Отправляем сообщение
            try:
                if photo_path and os.path.exists(photo_path):
                    # Отправляем фото с текстом
                    await self.client.send_file(entity, photo_path, caption=message_text if message_text else None)
                elif video_path and os.path.exists(video_path):
                    # Отправляем видео с текстом
                    await self.client.send_file(entity, video_path, caption=message_text if message_text else None)
                elif message_text:
                    # Отправляем только текст
                    await self.client.send_message(entity, message_text)
                else:
                    return {
                        'success': False,
                        'message': 'Не указан текст и не загружены файлы',
                        'blocked': False
                    }
                
                return {
                    'success': True,
                    'message': 'Сообщение отправлено успешно',
                    'blocked': False
                }
                
            except UserBannedInChannelError:
                return {
                    'success': False,
                    'message': 'Забанен в канале',
                    'blocked': True
                }
            except FloodWaitError as e:
                wait_time = e.seconds
                return {
                    'success': False,
                    'message': f'Flood wait: нужно подождать {wait_time} секунд',
                    'blocked': False
                }
            except Exception as e:
                error_msg = str(e).lower()
                if 'blocked' in error_msg or 'ban' in error_msg:
                    return {
                        'success': False,
                        'message': f'Заблокирован: {str(e)}',
                        'blocked': True
                    }
                return {
                    'success': False,
                    'message': f'Ошибка отправки: {str(e)}',
                    'blocked': False
                }
                
        except Exception as e:
            return {
                'success': False,
                'message': f'Исключение: {str(e)}',
                'blocked': False
            }
    
    def save_sending_report(self, results: List[Dict], report_file: str, sent_count: int, error_count: int, blocked_count: int, skipped_count: int):
        """
        Сохранение отчета о рассылке в Excel
        
        Args:
            results: Список результатов отправки
            report_file: Путь к файлу отчета
            sent_count: Количество успешно отправленных
            error_count: Количество ошибок
            blocked_count: Количество заблокированных
            skipped_count: Количество пропущенных
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sending Report"
        
        # Заголовки
        headers = ['ID', 'Название', 'Username', 'Участников', 'Статус', 'Сообщение', 'Время отправки', 'Ключевое слово']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Статистика в первой строке после заголовка
        ws.append(['', '', '', '', '', '', '', ''])
        ws.append(['СТАТИСТИКА', '', '', '', '', '', '', ''])
        ws.append(['✅ Отправлено успешно:', sent_count, '', '', '', '', '', ''])
        ws.append(['❌ Ошибки:', error_count, '', '', '', '', '', ''])
        ws.append(['🚫 Заблокировано:', blocked_count, '', '', '', '', '', ''])
        ws.append(['⏭️ Пропущено:', skipped_count, '', '', '', '', '', ''])
        ws.append(['', '', '', '', '', '', '', ''])
        ws.append(['ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ', '', '', '', '', '', '', ''])
        ws.append(headers)  # Повторяем заголовки
        
        # Данные
        for result in results:
            status = result.get('status', 'unknown')
            status_text = {
                'sent': '✅ Отправлено',
                'error': '❌ Ошибка',
                'blocked': '🚫 Заблокировано',
                'skipped': '⏭️ Пропущено'
            }.get(status, status)
            
            ws.append([
                result.get('id', 'N/A'),
                result.get('title', 'N/A'),
                result.get('username') or 'N/A',
                result.get('members_count', 'N/A'),
                status_text,
                result.get('message', ''),
                result.get('timestamp', 'N/A'),
                result.get('keyword', '')
            ])
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(report_file)
        print(f"✅ Отчет о рассылке сохранен: {report_file}")


async def main():
    """
    Основная функция для запуска поиска
    """
    # Попытка загрузить конфигурацию из config.py
    try:
        import config
        API_ID = config.API_ID
        API_HASH = config.API_HASH
        KEYWORDS = getattr(config, 'KEYWORDS', [])
        CITIES = getattr(config, 'CITIES', [])
        LIMIT_PER_KEYWORD = getattr(config, 'LIMIT_PER_KEYWORD', 50)
        USE_TRANSLITERATION = getattr(config, 'USE_TRANSLITERATION', True)
        USE_CITY_COMBINATIONS = getattr(config, 'USE_CITY_COMBINATIONS', True)
        SEARCH_DELAY = getattr(config, 'SEARCH_DELAY', 1.0)  # Задержка между запросами в секундах
        
        print("✅ Конфигурация загружена из config.py")
        
        # Генерируем поисковые запросы
        if USE_CITY_COMBINATIONS and CITIES:
            search_queries = TelegramSearcher.generate_search_queries(KEYWORDS, CITIES)
            print(f"📝 Сгенерировано {len(search_queries)} поисковых запросов из {len(KEYWORDS)} ключевых слов и {len(CITIES)} городов")
        elif USE_TRANSLITERATION:
            search_queries = TelegramSearcher.generate_search_queries(KEYWORDS)
            print(f"📝 Сгенерировано {len(search_queries)} поисковых запросов (с транслитерацией)")
        else:
            search_queries = KEYWORDS
            print(f"📝 Используется {len(search_queries)} ключевых слов (без комбинаций)")
        
    except ImportError:
        print("⚠️ Файл config.py не найден. Используются значения по умолчанию.")
        print("   Создайте config.py на основе config_example.py")
        # Значения по умолчанию (нужно заменить!)
        API_ID = 12345678  # ⚠️ Замените на ваш API ID
        API_HASH = 'your_api_hash_here'  # ⚠️ Замените на ваш API Hash
        KEYWORDS = ['python', 'programming', 'tech']
        CITIES = []
        LIMIT_PER_KEYWORD = 50
        USE_TRANSLITERATION = True
        USE_CITY_COMBINATIONS = False
        SEARCH_DELAY = 1.0
        search_queries = TelegramSearcher.generate_search_queries(KEYWORDS) if USE_TRANSLITERATION else KEYWORDS
    
    # Проверка корректности API credentials
    if API_ID == 12345678 or API_HASH == 'your_api_hash_here':
        print("\n❌ ОШИБКА: Необходимо настроить API_ID и API_HASH!")
        print("   1. Получите их на https://my.telegram.org")
        print("   2. Создайте config.py на основе config_example.py")
        print("   3. Заполните свои данные")
        return
    
    # Создание экземпляра поисковика
    try:
        searcher = TelegramSearcher(API_ID, API_HASH, search_delay=SEARCH_DELAY)
    except NameError:
        # Если SEARCH_DELAY не определен (старый config.py)
        searcher = TelegramSearcher(API_ID, API_HASH, search_delay=1.0)
    
    results = {'groups': [], 'channels': []}
    
    try:
        # Подключение
        await searcher.connect()
        
        # Поиск
        results = await searcher.search_channels_and_groups(search_queries, limit_per_keyword=LIMIT_PER_KEYWORD)
        
        # Сохранение в Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        groups_file = f'telegram_groups_{timestamp}.xlsx'
        channels_file = f'telegram_channels_{timestamp}.xlsx'
        
        searcher.save_to_excel(
            results['groups'],
            results['channels'],
            groups_file,
            channels_file
        )
        
        print(f"\n🎉 Готово! Результаты сохранены в файлы:")
        print(f"   📁 {groups_file}")
        print(f"   📁 {channels_file}")
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Поиск прерван пользователем (Ctrl+C)")
        print("💾 Сохраняю уже найденные результаты...")
        
        # Используем результаты из класса (они обновляются по мере поиска)
        saved_results = searcher.current_results
        
        if saved_results['groups'] or saved_results['channels']:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            groups_file = f'telegram_groups_interrupted_{timestamp}.xlsx'
            channels_file = f'telegram_channels_interrupted_{timestamp}.xlsx'
            
            searcher.save_to_excel(
                saved_results['groups'],
                saved_results['channels'],
                groups_file,
                channels_file
            )
            
            print(f"\n✅ Найденные результаты сохранены:")
            print(f"   📁 {groups_file} ({len(saved_results['groups'])} групп)")
            print(f"   📁 {channels_file} ({len(saved_results['channels'])} каналов)")
        else:
            print("⚠️ Результаты не найдены, сохранять нечего")
            print("   (Поиск был прерван до того, как что-то было найдено)")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        
        # Пытаемся сохранить результаты даже при ошибке
        saved_results = searcher.current_results
        if saved_results['groups'] or saved_results['channels']:
            print("\n💾 Пытаюсь сохранить найденные результаты...")
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                groups_file = f'telegram_groups_error_{timestamp}.xlsx'
                channels_file = f'telegram_channels_error_{timestamp}.xlsx'
                
                searcher.save_to_excel(
                    saved_results['groups'],
                    saved_results['channels'],
                    groups_file,
                    channels_file
                )
                print(f"✅ Результаты сохранены в файлы с префиксом 'error_'")
            except Exception as save_error:
                print(f"❌ Не удалось сохранить результаты: {save_error}")
    
    finally:
        await searcher.disconnect()


if __name__ == '__main__':
    # Запуск асинхронной функции
    asyncio.run(main())

