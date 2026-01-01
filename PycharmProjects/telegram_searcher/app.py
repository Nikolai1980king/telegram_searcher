"""
Flask веб-приложение для Telegram Searcher
Графический интерфейс для управления поиском групп и каналов
"""

from flask import Flask, render_template, request, jsonify, send_file, session
from flask_session import Session
import os
import json
import asyncio
import threading
from datetime import datetime
from pathlib import Path
import importlib.util
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Импорт основного класса поисковика
from telegram_searcher import TelegramSearcher

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = 'flask_session'
app.config['UPLOAD_FOLDER'] = 'results'
Session(app)

# Глобальные переменные для хранения состояния
search_tasks = {}  # {session_id: {'status': 'running'/'completed'/'error', 'results': {...}}}
search_configs = {}  # {session_id: {'keywords': [], 'cities': [], 'delay': 5.0}}
search_stop_flags = {}  # {session_id: threading.Event()} - флаги для остановки поиска
config_file_lock = threading.Lock()  # Блокировка для безопасной записи в config.py

# Переменные для проверки групп
check_groups_tasks = {}  # {session_id: {'status': 'running'/'completed'/'error', 'progress': {...}}}
check_groups_stop_flags = {}  # {session_id: threading.Event()}

# Переменные для обработки pending групп
process_pending_tasks = {}  # {session_id: {'status': 'running'/'completed'/'error', 'progress': {...}}}
process_pending_stop_flags = {}  # {session_id: threading.Event()}

# Создаем папку для результатов
os.makedirs('results', exist_ok=True)
os.makedirs('templates', exist_ok=True)
os.makedirs('static', exist_ok=True)


def get_session_id():
    """Получить или создать session ID"""
    if 'session_id' not in session:
        session['session_id'] = f"session_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    return session['session_id']


def parse_groups_from_text(text: str) -> List[Dict]:
    """
    Парсинг списка групп из текста
    Формат: каждая строка = один аккаунт (username или ID)
    
    Args:
        text: Текст со списком аккаунтов
        
    Returns:
        Список словарей с информацией о группах
    """
    groups = []
    lines = text.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Убираем @ если есть
        if line.startswith('@'):
            line = line[1:]
        
        # Пытаемся определить, это ID или username
        group_info = {}
        
        # Если это число - считаем ID
        try:
            group_id = int(line)
            group_info = {
                'id': group_id,
                'username': None,
                'title': f"ID: {group_id}"
            }
        except ValueError:
            # Это username
            group_info = {
                'id': None,
                'username': line,
                'title': line
            }
        
        groups.append(group_info)
    
    return groups


def save_config_to_file(keywords, cities, delay):
    """Сохранить конфигурацию в config.py"""
    # Используем блокировку для предотвращения одновременных записей
    with config_file_lock:
        try:
            import re
            import time
            config_path = 'config.py'
            
            # Небольшая задержка, чтобы избежать конфликтов при перезапуске Flask
            time.sleep(0.1)
            
            # Читаем текущий config.py (пробуем несколько раз, если файл пустой)
            content = None
            file_not_found = False
            for attempt in range(3):
                try:
                    with open(config_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    if content.strip() and 'KEYWORDS' in content:
                        break
                    elif attempt < 2:
                        app.logger.warning(f"⚠️ Попытка {attempt + 1}: config.py пустой, ждем...")
                        time.sleep(0.2)
                except FileNotFoundError:
                    file_not_found = True
                    if attempt < 2:
                        time.sleep(0.2)
                    break
            
            # Если файл не существует или пустой, создаем базовую структуру
            if file_not_found or not content or not content.strip() or 'KEYWORDS' not in content:
                if file_not_found:
                    app.logger.warning("⚠️ config.py не найден, создаем новый файл")
                else:
                    app.logger.warning("⚠️ config.py пустой или поврежден, восстанавливаем структуру")
                
                # Пытаемся сохранить API_ID и API_HASH из старого файла
                api_id = '27375139'
                api_hash = '66e1bc627b8dda02e2bb35ea44fde4cf'
                if content:
                    api_id_match = re.search(r'API_ID\s*=\s*(\d+)', content)
                    api_hash_match = re.search(r'API_HASH\s*=\s*["\']([^"\']+)["\']', content)
                    if api_id_match:
                        api_id = api_id_match.group(1)
                    if api_hash_match:
                        api_hash = api_hash_match.group(1)
                
                # Создаем новую структуру
                content = f'''"""
Файл конфигурации
"""

# Получите эти данные на https://my.telegram.org
API_ID = {api_id}
API_HASH = "{api_hash}"

# Ключевые слова для поиска
KEYWORDS = [
]

# Города для комбинаций с ключевыми словами
CITIES = [
]

# Максимальное количество результатов на одно ключевое слово
LIMIT_PER_KEYWORD = 50

# Использовать транслитерацию (русские -> английские буквы)
USE_TRANSLITERATION = True

# Использовать комбинации с городами
USE_CITY_COMBINATIONS = True

# Задержка между поисковыми запросами в секундах
SEARCH_DELAY = 5.0
'''
            
            # Формируем строку для KEYWORDS
            keywords_str = 'KEYWORDS = [\n'
            if keywords:
                for kw in keywords:
                    # Экранируем кавычки и обратные слеши
                    kw_escaped = kw.replace("\\", "\\\\").replace("'", "\\'").replace('"', '\\"')
                    keywords_str += f"    '{kw_escaped}',\n"
            keywords_str += ']'
            
            # Формируем строку для CITIES
            cities_str = 'CITIES = [\n'
            if cities:
                for city in cities:
                    # Экранируем кавычки и обратные слеши
                    city_escaped = city.replace("\\", "\\\\").replace('"', '\\"')
                    cities_str += f'    "{city_escaped}",\n'
            cities_str += ']'
            
            # Заменяем KEYWORDS - используем нежадный квантификатор для многострочных списков
            # Паттерн ищет KEYWORDS = [ и все до первой закрывающей скобки ], включая переносы строк
            keywords_pattern = r'KEYWORDS\s*=\s*\[.*?\]'
            if re.search(keywords_pattern, content, flags=re.DOTALL):
                content = re.sub(keywords_pattern, keywords_str, content, flags=re.DOTALL)
                app.logger.info(f"✅ KEYWORDS заменен в config.py: {len(keywords)} слов")
            else:
                app.logger.warning("⚠️ Не удалось найти KEYWORDS в config.py для замены")
            
            # Заменяем CITIES - аналогично
            cities_pattern = r'CITIES\s*=\s*\[.*?\]'
            if re.search(cities_pattern, content, flags=re.DOTALL):
                content = re.sub(cities_pattern, cities_str, content, flags=re.DOTALL)
                app.logger.info(f"✅ CITIES заменен в config.py: {len(cities)} городов")
            else:
                app.logger.warning("⚠️ Не удалось найти CITIES в config.py для замены")
            
            # Заменяем SEARCH_DELAY
            delay_pattern = r'SEARCH_DELAY\s*=\s*[\d.]+'
            content = re.sub(delay_pattern, f'SEARCH_DELAY = {delay}', content)
            
            # Сохраняем файл
            try:
                with open(config_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                # Принудительно сбрасываем буфер
                import os
                os.fsync(f.fileno()) if hasattr(f, 'fileno') else None
            except Exception as e:
                app.logger.error(f"❌ Ошибка записи в config.py: {e}")
                return False
            
            # Проверяем, что сохранение прошло успешно (пробуем несколько раз)
            for check_attempt in range(3):
                try:
                    time.sleep(0.05)  # Небольшая задержка для файловой системы
                    with open(config_path, 'r', encoding='utf-8') as f:
                        saved_content = f.read()
                    
                    if not saved_content.strip():
                        if check_attempt < 2:
                            app.logger.warning(f"⚠️ Проверка {check_attempt + 1}: файл пустой, повторяем...")
                            continue
                        else:
                            app.logger.error("❌ Файл остался пустым после сохранения!")
                            return False
                    
                    # Проверяем наличие KEYWORDS в файле
                    if 'KEYWORDS' not in saved_content:
                        app.logger.error("❌ KEYWORDS не найден в сохраненном файле!")
                        return False
                    
                    # Проверяем наличие ключевых слов
                    if keywords:
                        found_count = 0
                        for kw in keywords:
                            # Ищем слово в файле (может быть с экранированием)
                            kw_escaped = kw.replace("'", "\\'")
                            if kw in saved_content or kw_escaped in saved_content:
                                found_count += 1
                        
                        if found_count == 0:
                            app.logger.error(f"❌ Ключевые слова не найдены в сохраненном файле! Ожидалось: {keywords}")
                            return False
                        elif found_count < len(keywords):
                            app.logger.warning(f"⚠️ Найдено только {found_count} из {len(keywords)} ключевых слов")
                    
                    app.logger.info(f"✅ config.py успешно сохранен. KEYWORDS: {len(keywords)}, CITIES: {len(cities)}")
                    return True
                    
                except Exception as e:
                    if check_attempt < 2:
                        app.logger.warning(f"⚠️ Ошибка проверки сохранения (попытка {check_attempt + 1}): {e}")
                        time.sleep(0.1)
                    else:
                        app.logger.error(f"❌ Ошибка проверки сохранения: {e}")
                        return False
            
            return False
        except Exception as e:
            app.logger.error(f"Ошибка сохранения config.py: {e}")
            import traceback
            traceback.print_exc()
            return False


def run_search_async(session_id, keywords, cities, delay, api_id, api_hash):
    """Асинхронный запуск поиска в отдельном потоке"""
    try:
        app.logger.info(f"🚀 Запуск поиска: keywords={keywords}, cities={cities}, delay={delay}")
        
        # Проверяем входные данные
        if not keywords:
            search_tasks[session_id]['status'] = 'error'
            search_tasks[session_id]['message'] = 'Ошибка: не указаны ключевые слова'
            app.logger.error("❌ Ошибка: не указаны ключевые слова")
            return
        
        # Создаем флаг остановки
        stop_event = threading.Event()
        search_stop_flags[session_id] = stop_event
        
        # Создаем новый event loop для этого потока
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        # Обновляем статус
        search_tasks[session_id]['status'] = 'running'
        search_tasks[session_id]['message'] = 'Поиск запущен...'
        
        # Создаем экземпляр поисковика
        app.logger.info("📱 Создание экземпляра TelegramSearcher...")
        searcher = TelegramSearcher(api_id, api_hash, search_delay=delay)
        
        # Генерируем поисковые запросы
        app.logger.info(f"🔍 Генерация поисковых запросов из {len(keywords)} ключевых слов и {len(cities) if cities else 0} городов...")
        search_queries = TelegramSearcher.generate_search_queries(keywords, cities)
        app.logger.info(f"✅ Сгенерировано {len(search_queries)} поисковых запросов")
        
        # Запускаем поиск используя оригинальный метод класса
        async def search():
            try:
                await searcher.connect()
                
                # Используем оригинальный метод поиска
                search_tasks[session_id]['message'] = f'Поиск по {len(search_queries)} запросам...'
                
                # Модифицируем метод поиска для поддержки остановки
                # Создаем обертку, которая проверяет флаг остановки
                all_groups = []
                all_channels = []
                seen_ids = set()
                
                for i, keyword in enumerate(search_queries):
                    # Проверяем флаг остановки перед каждым запросом
                    if stop_event.is_set():
                        break
                    
                    try:
                        from telethon.tl.functions.contacts import SearchRequest
                        from telethon.tl.types import Channel, Chat
                        
                        results = await searcher.client(SearchRequest(
                            q=keyword,
                            limit=50
                        ))
                        
                        # Проверяем, есть ли результаты
                        found_in_query = 0
                        if hasattr(results, 'chats') and results.chats:
                            for result in results.chats:
                                if stop_event.is_set():
                                    break
                                
                                from telethon.tl.types import Channel, Chat
                                if not isinstance(result, (Channel, Chat)):
                                    continue
                                
                                entity_id = result.id
                                if entity_id in seen_ids:
                                    continue
                                seen_ids.add(entity_id)
                                
                                members_count = await searcher._get_members_count(result)
                                
                                entity_info = {
                                    'id': entity_id,
                                    'title': result.title,
                                    'username': getattr(result, 'username', None),
                                    'members_count': members_count,
                                    'keyword': keyword
                                }
                                
                                if isinstance(result, Channel):
                                    if result.broadcast:
                                        all_channels.append(entity_info)
                                        searcher.current_results['channels'].append(entity_info)
                                    else:
                                        all_groups.append(entity_info)
                                        searcher.current_results['groups'].append(entity_info)
                                elif isinstance(result, Chat):
                                    all_groups.append(entity_info)
                                    searcher.current_results['groups'].append(entity_info)
                                
                                found_in_query += 1
                        
                        # Обновляем прогресс с информацией о найденных
                        progress = (i + 1) / len(search_queries) * 100
                        found_total = len(all_groups) + len(all_channels)
                        if found_in_query > 0:
                            search_tasks[session_id]['message'] = f'Поиск... {i+1}/{len(search_queries)} ({progress:.1f}%) | Найдено: {found_total} | В этом запросе: {found_in_query}'
                        else:
                            search_tasks[session_id]['message'] = f'Поиск... {i+1}/{len(search_queries)} ({progress:.1f}%) | Найдено: {found_total}'
                        
                        # Обновляем прогресс
                        progress = (i + 1) / len(search_queries) * 100
                        found_total = len(all_groups) + len(all_channels)
                        search_tasks[session_id]['message'] = f'Поиск... {i+1}/{len(search_queries)} ({progress:.1f}%) | Найдено: {found_total}'
                        
                        # Задержка
                        if delay > 0 and not stop_event.is_set():
                            await asyncio.sleep(delay)
                            
                    except Exception as e:
                        error_msg = str(e)
                        # Обрабатываем flood wait
                        if "wait of" in error_msg and "seconds" in error_msg:
                            try:
                                wait_seconds = int(error_msg.split("wait of")[1].split("seconds")[0].strip())
                                wait_hours = wait_seconds / 3600
                                if wait_seconds > 3600:
                                    search_tasks[session_id]['message'] = f'⚠️ Flood wait: пропускаю "{keyword}" (ожидание ~{wait_hours:.1f}ч)'
                                    continue
                            except:
                                pass
                        
                        # Логируем другие ошибки, но продолжаем
                        if not stop_event.is_set():
                            # Не показываем каждую ошибку, только обновляем прогресс
                            continue
                        else:
                            break
                
                # Проверяем, был ли остановлен поиск
                if stop_event.is_set():
                    saved_results = searcher.current_results
                    if saved_results['groups'] or saved_results['channels']:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        groups_file = f'results/telegram_groups_stopped_{timestamp}.xlsx'
                        channels_file = f'results/telegram_channels_stopped_{timestamp}.xlsx'
                        
                        searcher.save_to_excel(
                            saved_results['groups'],
                            saved_results['channels'],
                            groups_file,
                            channels_file
                        )
                        
                        search_tasks[session_id]['status'] = 'stopped'
                        search_tasks[session_id]['message'] = f'Поиск остановлен. Найдено: {len(saved_results["groups"])} групп, {len(saved_results["channels"])} каналов'
                        search_tasks[session_id]['results'] = {
                            'groups_file': groups_file,
                            'channels_file': channels_file,
                            'groups_count': len(saved_results['groups']),
                            'channels_count': len(saved_results['channels'])
                        }
                    else:
                        search_tasks[session_id]['status'] = 'stopped'
                        search_tasks[session_id]['message'] = 'Поиск остановлен. Результаты не найдены'
                else:
                    # Поиск завершен успешно
                    results = {'groups': all_groups, 'channels': all_channels}
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    groups_file = f'results/telegram_groups_{timestamp}.xlsx'
                    channels_file = f'results/telegram_channels_{timestamp}.xlsx'
                    
                    searcher.save_to_excel(
                        results['groups'],
                        results['channels'],
                        groups_file,
                        channels_file
                    )
                    
                    search_tasks[session_id]['status'] = 'completed'
                    search_tasks[session_id]['message'] = f'Поиск завершен! Найдено: {len(results["groups"])} групп, {len(results["channels"])} каналов'
                    search_tasks[session_id]['results'] = {
                        'groups_file': groups_file,
                        'channels_file': channels_file,
                        'groups_count': len(results['groups']),
                        'channels_count': len(results['channels'])
                    }
                
                await searcher.disconnect()
                
            except KeyboardInterrupt:
                # Сохраняем результаты при прерывании
                saved_results = searcher.current_results
                if saved_results['groups'] or saved_results['channels']:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    groups_file = f'results/telegram_groups_interrupted_{timestamp}.xlsx'
                    channels_file = f'results/telegram_channels_interrupted_{timestamp}.xlsx'
                    
                    searcher.save_to_excel(
                        saved_results['groups'],
                        saved_results['channels'],
                        groups_file,
                        channels_file
                    )
                    
                    search_tasks[session_id]['status'] = 'interrupted'
                    search_tasks[session_id]['message'] = 'Поиск прерван, результаты сохранены'
                    search_tasks[session_id]['results'] = {
                        'groups_file': groups_file,
                        'channels_file': channels_file,
                        'groups_count': len(saved_results['groups']),
                        'channels_count': len(saved_results['channels'])
                    }
                
                await searcher.disconnect()
                
            except Exception as e:
                error_msg = f'Ошибка: {str(e)}'
                app.logger.error(f"❌ Ошибка в поиске: {e}")
                import traceback
                app.logger.error(traceback.format_exc())
                search_tasks[session_id]['status'] = 'error'
                search_tasks[session_id]['message'] = error_msg
                try:
                    await searcher.disconnect()
                except:
                    pass
        
        app.logger.info("🔄 Запуск event loop...")
        loop.run_until_complete(search())
        loop.close()
        app.logger.info("✅ Event loop завершен")
        
    except Exception as e:
        error_msg = f'Ошибка запуска: {str(e)}'
        app.logger.error(f"❌ Критическая ошибка запуска: {e}")
        import traceback
        app.logger.error(traceback.format_exc())
        search_tasks[session_id]['status'] = 'error'
        search_tasks[session_id]['message'] = error_msg


@app.route('/')
def index():
    """Главная страница"""
    session_id = get_session_id()
    
    # Загружаем данные из config.py
    try:
        import config as app_config
        keywords = getattr(app_config, 'KEYWORDS', [])
        cities = getattr(app_config, 'CITIES', [])
        delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
    except:
        keywords = []
        cities = []
        delay = 5.0
    
    # Инициализируем конфигурацию для сессии из config.py
    if session_id not in search_configs:
        search_configs[session_id] = {
            'keywords': keywords,
            'cities': cities,
            'delay': delay
        }
    else:
        # Обновляем из config.py если там есть изменения
        search_configs[session_id]['keywords'] = keywords
        search_configs[session_id]['cities'] = cities
        search_configs[session_id]['delay'] = delay
    
    # Инициализируем задачу поиска
    if session_id not in search_tasks:
        search_tasks[session_id] = {
            'status': 'idle',
            'message': 'Готов к запуску',
            'results': None
        }
    
    config = search_configs[session_id]
    return render_template('index.html', 
                         keywords=config['keywords'],
                         cities=config['cities'],
                         delay=config['delay'])

@app.route('/send_messages')
def send_messages_page():
    """Страница рассылки сообщений"""
    return render_template('send_messages.html')


@app.route('/api/add_keyword', methods=['POST'])
def add_keyword():
    """Добавить ключевое слово"""
    session_id = get_session_id()
    data = request.json
    keyword = data.get('keyword', '').strip()
    
    app.logger.info(f"➕ Добавление ключевого слова: '{keyword}' для сессии {session_id}")
    
    if not keyword:
        return jsonify({'success': False, 'message': 'Ключевое слово не может быть пустым'})
    
    # Если сессия не найдена, загружаем данные из config.py
    if session_id not in search_configs:
        app.logger.info(f"📝 Сессия не найдена, загружаем данные из config.py")
        try:
            import config as app_config
            keywords = getattr(app_config, 'KEYWORDS', [])
            cities = getattr(app_config, 'CITIES', [])
            delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
            search_configs[session_id] = {
                'keywords': keywords.copy(),
                'cities': cities.copy(),
                'delay': delay
            }
            app.logger.info(f"✅ Данные загружены из config.py: {len(keywords)} ключевых слов, {len(cities)} городов")
        except Exception as e:
            app.logger.error(f"❌ Ошибка загрузки из config.py: {e}")
            search_configs[session_id] = {'keywords': [], 'cities': [], 'delay': 5.0}
    
    if keyword not in search_configs[session_id]['keywords']:
        search_configs[session_id]['keywords'].append(keyword)
        app.logger.info(f"✅ Ключевое слово добавлено. Всего: {len(search_configs[session_id]['keywords'])}")
        app.logger.info(f"📋 Текущие ключевые слова: {search_configs[session_id]['keywords']}")
        
        # Сохраняем в config.py
        save_result = save_config_to_file(
            search_configs[session_id]['keywords'],
            search_configs[session_id]['cities'],
            search_configs[session_id]['delay']
        )
        if save_result:
            app.logger.info("💾 config.py успешно обновлен")
            # Проверяем, что данные действительно сохранились
            try:
                import config as test_config
                saved_keywords = getattr(test_config, 'KEYWORDS', [])
                app.logger.info(f"🔍 Проверка сохранения: в config.py найдено {len(saved_keywords)} ключевых слов")
                if len(saved_keywords) != len(search_configs[session_id]['keywords']):
                    app.logger.warning(f"⚠️ Несоответствие: в памяти {len(search_configs[session_id]['keywords'])}, в файле {len(saved_keywords)}")
            except Exception as e:
                app.logger.error(f"❌ Ошибка при проверке сохранения: {e}")
        else:
            app.logger.error("❌ Не удалось сохранить в config.py!")
        
        return jsonify({'success': True, 'keywords': search_configs[session_id]['keywords']})
    else:
        app.logger.warning(f"⚠️ Ключевое слово '{keyword}' уже существует")
        return jsonify({'success': False, 'message': 'Ключевое слово уже добавлено'})


@app.route('/api/remove_keyword', methods=['POST'])
def remove_keyword():
    """Удалить ключевое слово"""
    session_id = get_session_id()
    data = request.json
    keyword = data.get('keyword', '').strip()
    
    app.logger.info(f"➖ Удаление ключевого слова: '{keyword}' для сессии {session_id}")
    
    # Если сессия не найдена, загружаем данные из config.py
    if session_id not in search_configs:
        app.logger.info(f"📝 Сессия не найдена, загружаем данные из config.py")
        try:
            import config as app_config
            keywords = getattr(app_config, 'KEYWORDS', [])
            cities = getattr(app_config, 'CITIES', [])
            delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
            search_configs[session_id] = {
                'keywords': keywords.copy(),
                'cities': cities.copy(),
                'delay': delay
            }
            app.logger.info(f"✅ Данные загружены из config.py: {len(keywords)} ключевых слов, {len(cities)} городов")
        except Exception as e:
            app.logger.error(f"❌ Ошибка загрузки из config.py: {e}")
            search_configs[session_id] = {'keywords': [], 'cities': [], 'delay': 5.0}
    
    app.logger.info(f"📋 Текущие ключевые слова: {search_configs[session_id]['keywords']}")
    
    if keyword in search_configs[session_id]['keywords']:
        search_configs[session_id]['keywords'].remove(keyword)
        app.logger.info(f"✅ Ключевое слово удалено. Осталось: {len(search_configs[session_id]['keywords'])}")
        
        # Сохраняем в config.py
        if save_config_to_file(
            search_configs[session_id]['keywords'],
            search_configs[session_id]['cities'],
            search_configs[session_id]['delay']
        ):
            app.logger.info("💾 config.py обновлен")
        else:
            app.logger.warning("⚠️ Не удалось сохранить в config.py")
        
        return jsonify({'success': True, 'keywords': search_configs[session_id]['keywords']})
    else:
        app.logger.warning(f"⚠️ Ключевое слово '{keyword}' не найдено в списке")
        return jsonify({'success': False, 'message': f'Ключевое слово "{keyword}" не найдено'})


@app.route('/api/add_city', methods=['POST'])
def add_city():
    """Добавить город"""
    session_id = get_session_id()
    data = request.json
    city = data.get('city', '').strip()
    
    app.logger.info(f"➕ Добавление города: '{city}' для сессии {session_id}")
    
    if not city:
        return jsonify({'success': False, 'message': 'Город не может быть пустым'})
    
    # Если сессия не найдена, загружаем данные из config.py
    if session_id not in search_configs:
        app.logger.info(f"📝 Сессия не найдена, загружаем данные из config.py")
        try:
            import config as app_config
            keywords = getattr(app_config, 'KEYWORDS', [])
            cities = getattr(app_config, 'CITIES', [])
            delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
            search_configs[session_id] = {
                'keywords': keywords.copy(),
                'cities': cities.copy(),
                'delay': delay
            }
            app.logger.info(f"✅ Данные загружены из config.py: {len(keywords)} ключевых слов, {len(cities)} городов")
        except Exception as e:
            app.logger.error(f"❌ Ошибка загрузки из config.py: {e}")
            search_configs[session_id] = {'keywords': [], 'cities': [], 'delay': 5.0}
    
    if city not in search_configs[session_id]['cities']:
        search_configs[session_id]['cities'].append(city)
        app.logger.info(f"✅ Город добавлен. Всего: {len(search_configs[session_id]['cities'])}")
        app.logger.info(f"📋 Текущие города: {search_configs[session_id]['cities']}")
        app.logger.info(f"📋 Текущие ключевые слова: {search_configs[session_id]['keywords']}")
        
        # Сохраняем в config.py (с текущими ключевыми словами!)
        save_result = save_config_to_file(
            search_configs[session_id]['keywords'],
            search_configs[session_id]['cities'],
            search_configs[session_id]['delay']
        )
        if save_result:
            app.logger.info("💾 config.py успешно обновлен")
        else:
            app.logger.error("❌ Не удалось сохранить в config.py!")
        
        return jsonify({'success': True, 'cities': search_configs[session_id]['cities']})
    else:
        app.logger.warning(f"⚠️ Город '{city}' уже существует")
        return jsonify({'success': False, 'message': 'Город уже добавлен'})


@app.route('/api/remove_city', methods=['POST'])
def remove_city():
    """Удалить город"""
    session_id = get_session_id()
    data = request.json
    city = data.get('city', '').strip()
    
    app.logger.info(f"➖ Удаление города: '{city}' для сессии {session_id}")
    
    # Если сессия не найдена, загружаем данные из config.py
    if session_id not in search_configs:
        app.logger.info(f"📝 Сессия не найдена, загружаем данные из config.py")
        try:
            import config as app_config
            keywords = getattr(app_config, 'KEYWORDS', [])
            cities = getattr(app_config, 'CITIES', [])
            delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
            search_configs[session_id] = {
                'keywords': keywords.copy(),
                'cities': cities.copy(),
                'delay': delay
            }
            app.logger.info(f"✅ Данные загружены из config.py: {len(keywords)} ключевых слов, {len(cities)} городов")
        except Exception as e:
            app.logger.error(f"❌ Ошибка загрузки из config.py: {e}")
            search_configs[session_id] = {'keywords': [], 'cities': [], 'delay': 5.0}
    
    app.logger.info(f"📋 Текущие города: {search_configs[session_id]['cities']}")
    
    if city in search_configs[session_id]['cities']:
        search_configs[session_id]['cities'].remove(city)
        app.logger.info(f"✅ Город удален. Осталось: {len(search_configs[session_id]['cities'])}")
        
        # Сохраняем в config.py
        if save_config_to_file(
            search_configs[session_id]['keywords'],
            search_configs[session_id]['cities'],
            search_configs[session_id]['delay']
        ):
            app.logger.info("💾 config.py обновлен")
        else:
            app.logger.warning("⚠️ Не удалось сохранить в config.py")
        
        return jsonify({'success': True, 'cities': search_configs[session_id]['cities']})
    else:
        app.logger.warning(f"⚠️ Город '{city}' не найден в списке")
        return jsonify({'success': False, 'message': f'Город "{city}" не найден'})


@app.route('/api/set_delay', methods=['POST'])
def set_delay():
    """Установить задержку"""
    session_id = get_session_id()
    data = request.json
    delay = float(data.get('delay', 5.0))
    
    app.logger.info(f"⏱️ Установка задержки: {delay} секунд для сессии {session_id}")
    
    if delay < 0:
        return jsonify({'success': False, 'message': 'Задержка не может быть отрицательной'})
    
    # Если сессия не найдена, загружаем данные из config.py
    if session_id not in search_configs:
        app.logger.info(f"📝 Сессия не найдена, загружаем данные из config.py")
        try:
            import config as app_config
            keywords = getattr(app_config, 'KEYWORDS', [])
            cities = getattr(app_config, 'CITIES', [])
            current_delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
            search_configs[session_id] = {
                'keywords': keywords.copy(),
                'cities': cities.copy(),
                'delay': current_delay
            }
            app.logger.info(f"✅ Данные загружены из config.py: {len(keywords)} ключевых слов, {len(cities)} городов")
        except Exception as e:
            app.logger.error(f"❌ Ошибка загрузки из config.py: {e}")
            search_configs[session_id] = {'keywords': [], 'cities': [], 'delay': 5.0}
    
    search_configs[session_id]['delay'] = delay
    app.logger.info(f"📋 Текущие ключевые слова: {search_configs[session_id]['keywords']}")
    app.logger.info(f"📋 Текущие города: {search_configs[session_id]['cities']}")
    
    # Сохраняем в config.py (с текущими ключевыми словами и городами!)
    save_result = save_config_to_file(
        search_configs[session_id]['keywords'],
        search_configs[session_id]['cities'],
        delay
    )
    if save_result:
        app.logger.info("💾 config.py успешно обновлен")
    else:
        app.logger.error("❌ Не удалось сохранить в config.py!")
    
    return jsonify({'success': True, 'delay': delay})


@app.route('/api/start_search', methods=['POST'])
def start_search():
    """Запустить поиск"""
    session_id = get_session_id()
    app.logger.info(f"🔍 Запрос на запуск поиска от сессии {session_id}")
    
    # Если сессия не найдена, загружаем данные из config.py
    if session_id not in search_configs:
        app.logger.info(f"📝 Сессия не найдена, загружаем данные из config.py")
        try:
            import config as app_config
            keywords = getattr(app_config, 'KEYWORDS', [])
            cities = getattr(app_config, 'CITIES', [])
            delay = getattr(app_config, 'SEARCH_DELAY', 5.0)
            search_configs[session_id] = {
                'keywords': keywords.copy(),
                'cities': cities.copy(),
                'delay': delay
            }
            app.logger.info(f"✅ Данные загружены из config.py: {len(keywords)} ключевых слов, {len(cities)} городов")
        except Exception as e:
            app.logger.error(f"❌ Ошибка загрузки из config.py: {e}")
            search_configs[session_id] = {'keywords': [], 'cities': [], 'delay': 5.0}
    
    config = search_configs[session_id]
    app.logger.info(f"📋 Конфигурация: keywords={len(config.get('keywords', []))}, cities={len(config.get('cities', []))}, delay={config.get('delay', 5.0)}")
    
    if not config.get('keywords'):
        app.logger.warning("⚠️ Ключевые слова не указаны")
        return jsonify({'success': False, 'message': 'Добавьте хотя бы одно ключевое слово'})
    
    # Проверяем API credentials
    try:
        import config as app_config
        api_id = app_config.API_ID
        api_hash = app_config.API_HASH
        app.logger.info("✅ API credentials загружены")
    except Exception as e:
        app.logger.error(f"❌ Ошибка загрузки API credentials: {e}")
        return jsonify({'success': False, 'message': 'API_ID и API_HASH не настроены в config.py'})
    
    # Проверяем, не запущен ли уже поиск
    if session_id in search_tasks and search_tasks[session_id]['status'] == 'running':
        app.logger.warning("⚠️ Поиск уже запущен")
        return jsonify({'success': False, 'message': 'Поиск уже запущен'})
    
    # Инициализируем задачу
    search_tasks[session_id] = {
        'status': 'starting',
        'message': 'Запуск поиска...',
        'results': None
    }
    
    # Запускаем поиск в отдельном потоке
    app.logger.info("🚀 Создание потока для поиска...")
    thread = threading.Thread(
        target=run_search_async,
        args=(session_id, config['keywords'], config.get('cities', []), config.get('delay', 5.0), api_id, api_hash)
    )
    thread.daemon = True
    thread.start()
    app.logger.info("✅ Поток запущен")
    
    return jsonify({'success': True, 'message': 'Поиск запущен'})


@app.route('/api/stop_search', methods=['POST'])
def stop_search():
    """Остановить поиск"""
    session_id = get_session_id()
    
    if session_id in search_stop_flags:
        search_stop_flags[session_id].set()
        search_tasks[session_id]['message'] = 'Остановка поиска...'
        return jsonify({'success': True, 'message': 'Команда остановки отправлена'})
    
    return jsonify({'success': False, 'message': 'Поиск не запущен'})


@app.route('/api/status', methods=['GET'])
def get_status():
    """Получить статус поиска"""
    session_id = get_session_id()
    
    if session_id not in search_tasks:
        return jsonify({
            'status': 'idle',
            'message': 'Готов к запуску',
            'results': None
        })
    
    task = search_tasks[session_id]
    return jsonify({
        'status': task['status'],
        'message': task.get('message', ''),
        'results': task.get('results')
    })


@app.route('/api/download/<filename>')
def download_file(filename):
    """Скачать файл результата"""
    # Безопасность: проверяем, что filename не содержит путь
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'error': 'Недопустимое имя файла'}), 400
    
    file_path = os.path.join('results', filename)
    
    if os.path.exists(file_path) and os.path.isfile(file_path):
        try:
            return send_file(file_path, as_attachment=True, download_name=filename)
        except Exception as e:
            app.logger.error(f"Ошибка при скачивании файла {filename}: {e}")
            return jsonify({'error': f'Ошибка при скачивании: {str(e)}'}), 500
    else:
        app.logger.warning(f"Файл не найден: {file_path}")
        return jsonify({'error': 'Файл не найден'}), 404


@app.route('/api/get_files', methods=['GET'])
def get_files():
    """Получить список доступных файлов результатов"""
    files = []
    results_dir = Path('results')
    
    if results_dir.exists():
        for file in results_dir.glob('*.xlsx'):
            files.append({
                'name': file.name,
                'size': file.stat().st_size,
                'modified': datetime.fromtimestamp(file.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')
            })
    
    # Сортируем по дате изменения (новые первыми)
    files.sort(key=lambda x: x['modified'], reverse=True)
    
    # Также возвращаем простой список имен для обратной совместимости
    file_names = [f['name'] for f in files]
    
    return jsonify({'files': files, 'file_names': file_names})


def run_check_groups_async(session_id, filename, api_id, api_hash):
    """Запуск проверки групп в отдельном потоке"""
    def run():
        try:
            check_groups_tasks[session_id] = {
                'status': 'running',
                'progress': {'current': 0, 'total': 0, 'message': 'Инициализация...'},
                'result_file': None
            }
            
            # Создаем новый event loop для этого потока
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            async def check_groups():
                try:
                    # Инициализация клиента
                    searcher = TelegramSearcher(api_id, api_hash, search_delay=2.0)
                    await searcher.connect()
                    
                    # Читаем группы из файла
                    filepath = os.path.join('results', filename) if not os.path.isabs(filename) else filename
                    if not os.path.exists(filepath):
                        filepath = filename  # Пробуем прямой путь
                    
                    app.logger.info(f"📖 Чтение групп из файла: {filepath}")
                    groups = TelegramSearcher.read_groups_from_excel(filepath)
                    
                    if not groups:
                        check_groups_tasks[session_id] = {
                            'status': 'error',
                            'progress': {'current': 0, 'total': 0, 'message': 'Группы не найдены в файле'},
                            'result_file': None
                        }
                        await searcher.disconnect()
                        return
                    
                    app.logger.info(f"✅ Найдено {len(groups)} групп для проверки")
                    
                    # Обновляем прогресс
                    check_groups_tasks[session_id]['progress'] = {
                        'current': 0,
                        'total': len(groups),
                        'message': f'Начало проверки {len(groups)} групп...'
                    }
                    
                    checked_groups = []
                    ready_count = 0
                    pending_count = 0
                    unavailable_count = 0
                    
                    stop_event = check_groups_stop_flags.get(session_id)
                    
                    # Проверяем каждую группу
                    for i, group in enumerate(groups):
                        if stop_event and stop_event.is_set():
                            check_groups_tasks[session_id]['status'] = 'stopped'
                            check_groups_tasks[session_id]['progress']['message'] = 'Проверка остановлена'
                            break
                        
                        group_title = group.get('title', f"ID: {group.get('id', 'N/A')}")
                        check_groups_tasks[session_id]['progress'] = {
                            'current': i + 1,
                            'total': len(groups),
                            'message': f'Проверяю: {group_title}',
                            'current_group': group_title
                        }
                        
                        app.logger.info(f"🔍 Проверка группы {i+1}/{len(groups)}: {group_title}")
                        
                        # Проверяем доступ
                        result = await searcher.check_group_access(group, stop_event)
                        
                        group['check_status'] = result.get('status', 'error')
                        group['check_message'] = result.get('message', '')
                        group['check_action'] = result.get('action_taken', 'none')
                        checked_groups.append(group)
                        
                        # Подсчитываем статистику
                        if result.get('status') == 'ready':
                            ready_count += 1
                        elif result.get('status') == 'pending':
                            pending_count += 1
                        else:
                            unavailable_count += 1
                        
                        # Задержка между проверками
                        if i < len(groups) - 1:  # Не ждем после последней группы
                            await asyncio.sleep(2.0)  # Задержка 2 секунды
                    
                    # Сохраняем результаты в два файла
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    ready_filename = f'ready_groups_{timestamp}.xlsx'
                    pending_filename = f'pending_groups_{timestamp}.xlsx'
                    ready_file = os.path.join('results', ready_filename)
                    pending_file = os.path.join('results', pending_filename)
                    
                    saved_ready, saved_pending = searcher.save_check_results(
                        checked_groups, 
                        ready_file, 
                        pending_file
                    )
                    
                    check_groups_tasks[session_id] = {
                        'status': 'completed',
                        'progress': {
                            'current': len(checked_groups),
                            'total': len(groups),
                            'message': 'Проверка завершена'
                        },
                        'ready_file': ready_filename if saved_ready > 0 else None,
                        'pending_file': pending_filename if saved_pending > 0 else None,
                        'ready_count': ready_count,
                        'pending_count': pending_count,
                        'unavailable_count': unavailable_count
                    }
                    
                    await searcher.disconnect()
                    app.logger.info(f"✅ Проверка завершена. Готовых: {ready_count}, Требуют действий: {pending_count}, Недоступных: {unavailable_count}")
                    
                except Exception as e:
                    app.logger.error(f"❌ Ошибка при проверке групп: {e}", exc_info=True)
                    check_groups_tasks[session_id] = {
                        'status': 'error',
                        'progress': {'current': 0, 'total': 0, 'message': f'Ошибка: {str(e)}'},
                        'result_file': None
                    }
            
            loop.run_until_complete(check_groups())
            loop.close()
            
        except Exception as e:
            app.logger.error(f"❌ Критическая ошибка в потоке проверки: {e}", exc_info=True)
            check_groups_tasks[session_id] = {
                'status': 'error',
                'progress': {'current': 0, 'total': 0, 'message': f'Критическая ошибка: {str(e)}'},
                'result_file': None
            }
    
    thread = threading.Thread(target=run)
    thread.daemon = True
    thread.start()


@app.route('/api/check_groups', methods=['POST'])
def check_groups():
    """Запустить проверку групп"""
    session_id = get_session_id()
    data = request.json
    filename = data.get('filename', '').strip()
    
    app.logger.info(f"🔍 Запрос на проверку групп от сессии {session_id}, файл: {filename}")
    
    if not filename:
        return jsonify({'success': False, 'message': 'Не указан файл'})
    
    # Проверяем API credentials
    try:
        import config as app_config
        api_id = app_config.API_ID
        api_hash = app_config.API_HASH
    except Exception as e:
        app.logger.error(f"❌ Ошибка загрузки API credentials: {e}")
        return jsonify({'success': False, 'message': 'Ошибка загрузки API credentials'})
    
    # Проверяем, не запущена ли уже проверка
    if session_id in check_groups_tasks and check_groups_tasks[session_id]['status'] == 'running':
        return jsonify({'success': False, 'message': 'Проверка уже запущена'})
    
    # Создаем флаг остановки
    check_groups_stop_flags[session_id] = threading.Event()
    
    # Запускаем проверку в отдельном потоке
    app.logger.info("🚀 Создание потока для проверки групп...")
    run_check_groups_async(session_id, filename, api_id, api_hash)
    app.logger.info("✅ Поток запущен")
    
    return jsonify({'success': True, 'message': 'Проверка запущена'})


@app.route('/api/stop_check_groups', methods=['POST'])
def stop_check_groups():
    """Остановить проверку групп"""
    session_id = get_session_id()
    
    if session_id in check_groups_stop_flags:
        check_groups_stop_flags[session_id].set()
        if session_id in check_groups_tasks:
            check_groups_tasks[session_id]['status'] = 'stopped'
            check_groups_tasks[session_id]['progress']['message'] = 'Остановка проверки...'
        return jsonify({'success': True, 'message': 'Команда остановки отправлена'})
    
    return jsonify({'success': False, 'message': 'Проверка не запущена'})


@app.route('/api/check_groups_status', methods=['GET'])
def check_groups_status():
    """Получить статус проверки групп"""
    session_id = get_session_id()
    
    if session_id not in check_groups_tasks:
        return jsonify({
            'status': 'idle',
            'message': 'Проверка не запущена',
            'current': 0,
            'total': 0
        })
    
    task = check_groups_tasks[session_id]
    progress = task.get('progress', {})
    
    response = {
        'status': task.get('status', 'idle'),
        'current': progress.get('current', 0),
        'total': progress.get('total', 0),
        'message': progress.get('message', ''),
        'current_group': progress.get('current_group', '')
    }
    
    if task.get('status') == 'completed':
        response['ready_file'] = task.get('ready_file')
        response['pending_file'] = task.get('pending_file')
        response['ready_count'] = task.get('ready_count', 0)
        response['pending_count'] = task.get('pending_count', 0)
        response['unavailable_count'] = task.get('unavailable_count', 0)
    
    return jsonify(response)


def run_process_pending_async(session_id, filename, api_id, api_hash):
    """Запуск обработки pending групп в отдельном потоке"""
    def run():
        try:
            process_pending_tasks[session_id] = {
                'status': 'running',
                'progress': {'current': 0, 'total': 0, 'message': 'Инициализация...'},
                'new_ready_file': None,
                'updated_pending_file': None
            }
            
            # Создаем новый event loop для этого потока
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            async def process_pending():
                try:
                    # Инициализация клиента
                    searcher = TelegramSearcher(api_id, api_hash, search_delay=2.0)
                    await searcher.connect()
                    
                    # Читаем pending группы из файла
                    filepath = os.path.join('results', filename) if not os.path.isabs(filename) else filename
                    if not os.path.exists(filepath):
                        filepath = filename
                    
                    app.logger.info(f"📖 Чтение pending групп из файла: {filepath}")
                    pending_groups = TelegramSearcher.read_groups_from_excel(filepath)
                    
                    if not pending_groups:
                        process_pending_tasks[session_id] = {
                            'status': 'error',
                            'progress': {'current': 0, 'total': 0, 'message': 'Группы не найдены в файле'},
                            'new_ready_file': None,
                            'updated_pending_file': None
                        }
                        await searcher.disconnect()
                        return
                    
                    app.logger.info(f"✅ Найдено {len(pending_groups)} pending групп для обработки")
                    
                    # Обновляем прогресс
                    process_pending_tasks[session_id]['progress'] = {
                        'current': 0,
                        'total': len(pending_groups),
                        'message': f'Начало обработки {len(pending_groups)} групп...'
                    }
                    
                    stop_event = process_pending_stop_flags.get(session_id)
                    
                    # Функция для обновления прогресса
                    def update_progress(current, total, message, current_group):
                        process_pending_tasks[session_id]['progress'] = {
                            'current': current,
                            'total': total,
                            'message': message,
                            'current_group': current_group
                        }
                    
                    # Обрабатываем pending группы
                    results = await searcher.process_pending_groups(pending_groups, stop_event, update_progress)
                    
                    # Сохраняем результаты
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    
                    new_ready_file = None
                    updated_pending_file = None
                    
                    # Сохраняем новые готовые группы
                    if results['ready_groups']:
                        new_ready_filename = f'new_ready_groups_{timestamp}.xlsx'
                        new_ready_file_path = os.path.join('results', new_ready_filename)
                        # Используем существующую функцию сохранения (готовые + пустой pending)
                        searcher.save_check_results(
                            results['ready_groups'],
                            new_ready_file_path,
                            os.path.join('results', 'temp_pending.xlsx')  # Временный файл, потом удалим
                        )
                        # Удаляем временный файл
                        temp_file = os.path.join('results', 'temp_pending.xlsx')
                        if os.path.exists(temp_file):
                            os.remove(temp_file)
                        new_ready_file = new_ready_filename
                    
                    # Сохраняем обновленный pending файл
                    if results['still_pending']:
                        updated_pending_filename = f'updated_pending_groups_{timestamp}.xlsx'
                        updated_pending_file_path = os.path.join('results', updated_pending_filename)
                        # Создаем файл с pending группами вручную
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Pending Groups"
                        headers = ['ID', 'Название', 'Username', 'Участников', 'Статус', 'Сообщение', 'Действие', 'Ключевое слово', 'Родительская группа']
                        ws.append(headers)
                        header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                        for group in results['still_pending']:
                            status_text = '⏳ Требует действий'
                            if group.get('check_status') == 'error':
                                status_text = '⚠️ Ошибка'
                            ws.append([
                                group.get('id', 'N/A'),
                                group.get('title', 'N/A'),
                                group.get('username') or 'N/A',
                                group.get('members_count', 'N/A'),
                                status_text,
                                group.get('check_message', ''),
                                group.get('check_action', ''),
                                group.get('keyword', ''),
                                group.get('parent_group', 'N/A')  # Для тем форумов
                            ])
                        # Автоматическая ширина колонок
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        wb.save(updated_pending_file_path)
                        updated_pending_file = updated_pending_filename
                    
                    process_pending_tasks[session_id] = {
                        'status': 'completed',
                        'progress': {
                            'current': len(pending_groups),
                            'total': len(pending_groups),
                            'message': 'Обработка завершена'
                        },
                        'new_ready_file': new_ready_file,
                        'updated_pending_file': updated_pending_file,
                        'new_ready_count': len(results['ready_groups']),
                        'still_pending_count': len(results['still_pending'])
                    }
                    
                    await searcher.disconnect()
                    app.logger.info(f"✅ Обработка завершена. Новых готовых: {len(results['ready_groups'])}, Все еще pending: {len(results['still_pending'])}")
                    
                except Exception as e:
                    app.logger.error(f"❌ Ошибка при обработке pending групп: {e}", exc_info=True)
                    process_pending_tasks[session_id] = {
                        'status': 'error',
                        'progress': {'current': 0, 'total': 0, 'message': f'Ошибка: {str(e)}'},
                        'new_ready_file': None,
                        'updated_pending_file': None
                    }
            
            loop.run_until_complete(process_pending())
            loop.close()
            
        except Exception as e:
            app.logger.error(f"❌ Критическая ошибка в потоке обработки pending: {e}", exc_info=True)
            process_pending_tasks[session_id] = {
                'status': 'error',
                'progress': {'current': 0, 'total': 0, 'message': f'Критическая ошибка: {str(e)}'},
                'new_ready_file': None,
                'updated_pending_file': None
            }
    
    thread = threading.Thread(target=run)
    thread.daemon = True
    thread.start()


@app.route('/api/process_pending_groups', methods=['POST'])
def process_pending_groups():
    """Запустить обработку pending групп"""
    session_id = get_session_id()
    data = request.json
    filename = data.get('filename', '').strip()
    
    app.logger.info(f"🔄 Запрос на обработку pending групп от сессии {session_id}, файл: {filename}")
    
    if not filename:
        return jsonify({'success': False, 'message': 'Не указан файл'})
    
    if 'pending' not in filename:
        return jsonify({'success': False, 'message': 'Выберите файл с pending группами'})
    
    # Проверяем API credentials
    try:
        import config as app_config
        api_id = app_config.API_ID
        api_hash = app_config.API_HASH
    except Exception as e:
        app.logger.error(f"❌ Ошибка загрузки API credentials: {e}")
        return jsonify({'success': False, 'message': 'Ошибка загрузки API credentials'})
    
    # Проверяем, не запущена ли уже обработка
    if session_id in process_pending_tasks and process_pending_tasks[session_id]['status'] == 'running':
        return jsonify({'success': False, 'message': 'Обработка уже запущена'})
    
    # Создаем флаг остановки
    process_pending_stop_flags[session_id] = threading.Event()
    
    # Запускаем обработку в отдельном потоке
    app.logger.info("🚀 Создание потока для обработки pending групп...")
    run_process_pending_async(session_id, filename, api_id, api_hash)
    app.logger.info("✅ Поток запущен")
    
    return jsonify({'success': True, 'message': 'Обработка запущена'})


@app.route('/api/stop_process_pending', methods=['POST'])
def stop_process_pending():
    """Остановить обработку pending групп"""
    session_id = get_session_id()
    
    if session_id in process_pending_stop_flags:
        process_pending_stop_flags[session_id].set()
        if session_id in process_pending_tasks:
            process_pending_tasks[session_id]['status'] = 'stopped'
            process_pending_tasks[session_id]['progress']['message'] = 'Остановка обработки...'
        return jsonify({'success': True, 'message': 'Команда остановки отправлена'})
    
    return jsonify({'success': False, 'message': 'Обработка не запущена'})


@app.route('/api/process_pending_status', methods=['GET'])
def process_pending_status():
    """Получить статус обработки pending групп"""
    session_id = get_session_id()
    
    if session_id not in process_pending_tasks:
        return jsonify({
            'status': 'idle',
            'message': 'Обработка не запущена',
            'current': 0,
            'total': 0
        })
    
    task = process_pending_tasks[session_id]
    progress = task.get('progress', {})
    
    response = {
        'status': task.get('status', 'idle'),
        'current': progress.get('current', 0),
        'total': progress.get('total', 0),
        'message': progress.get('message', ''),
        'current_group': progress.get('current_group', '')
    }
    
    if task.get('status') == 'completed':
        response['new_ready_file'] = task.get('new_ready_file')
        response['updated_pending_file'] = task.get('updated_pending_file')
        response['new_ready_count'] = task.get('new_ready_count', 0)
        response['still_pending_count'] = task.get('still_pending_count', 0)
    
    return jsonify(response)


@app.route('/api/merge_ready_groups', methods=['POST'])
def merge_ready_groups():
    """Объединить все файлы ready_groups в один"""
    try:
        results_dir = Path('results')
        if not results_dir.exists():
            return jsonify({'success': False, 'message': 'Папка results не найдена'})
        
        # Находим все файлы ready_groups
        ready_files = list(results_dir.glob('ready_groups_*.xlsx'))
        ready_files += list(results_dir.glob('new_ready_groups_*.xlsx'))
        
        if not ready_files:
            return jsonify({'success': False, 'message': 'Файлы ready_groups не найдены'})
        
        app.logger.info(f"📋 Найдено {len(ready_files)} файлов ready_groups для объединения")
        
        # Читаем все группы из всех файлов
        all_groups = []
        seen_ids = set()  # Для удаления дубликатов
        
        for file_path in ready_files:
            try:
                groups = TelegramSearcher.read_groups_from_excel(str(file_path))
                for group in groups:
                    group_id = group.get('id')
                    # Проверяем на дубликаты по ID
                    if group_id and group_id not in seen_ids:
                        seen_ids.add(group_id)
                        all_groups.append(group)
                    elif not group_id:
                        # Если нет ID, добавляем все равно (может быть тема форума)
                        all_groups.append(group)
                app.logger.info(f"  ✅ Из {file_path.name}: {len(groups)} групп")
            except Exception as e:
                app.logger.error(f"  ❌ Ошибка чтения {file_path.name}: {e}")
        
        if not all_groups:
            return jsonify({'success': False, 'message': 'Не удалось прочитать группы из файлов'})
        
        # Сохраняем объединенный файл
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        result_filename = f'all_ready_groups_{timestamp}.xlsx'
        result_file = os.path.join('results', result_filename)
        
        # Создаем Excel файл с готовыми группами
        wb = Workbook()
        ws = wb.active
        ws.title = "All Ready Groups"
        
        headers = ['ID', 'Название', 'Username', 'Участников', 'Статус', 'Сообщение', 'Действие', 'Ключевое слово', 'Родительская группа']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for group in all_groups:
            ws.append([
                group.get('id', 'N/A'),
                group.get('title', 'N/A'),
                group.get('username') or 'N/A',
                group.get('members_count', 'N/A'),
                '✅ Готово к рассылке',
                group.get('check_message', '') or 'Готово к рассылке',
                group.get('check_action', '') or 'none',
                group.get('keyword', ''),
                group.get('parent_group', 'N/A')
            ])
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(result_file)
        
        app.logger.info(f"✅ Объединено {len(ready_files)} файлов, всего {len(all_groups)} групп в {result_filename}")
        
        return jsonify({
            'success': True,
            'result_file': result_filename,
            'files_count': len(ready_files),
            'total_groups': len(all_groups)
        })
        
    except Exception as e:
        app.logger.error(f"❌ Ошибка при объединении файлов: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

@app.route('/merge_files')
def merge_files_page():
    """Страница объединения файлов с удалением дубликатов"""
    return render_template('merge_files.html')

@app.route('/api/merge_uploaded_files', methods=['POST'])
def merge_uploaded_files():
    """Объединить загруженные файлы с удалением дубликатов"""
    try:
        if 'files' not in request.files:
            return jsonify({'success': False, 'message': 'Не выбраны файлы для загрузки'})
        
        files = request.files.getlist('files')
        if not files or all(f.filename == '' for f in files):
            return jsonify({'success': False, 'message': 'Не выбраны файлы для загрузки'})
        
        app.logger.info(f"📋 Получено {len(files)} файлов для объединения")
        
        # Сохраняем загруженные файлы временно
        uploaded_files = []
        uploads_dir = Path('uploads')
        uploads_dir.mkdir(exist_ok=True)
        
        for file in files:
            if file.filename and file.filename.endswith(('.xlsx', '.xls')):
                filename = f"merge_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
                filepath = uploads_dir / filename
                file.save(str(filepath))
                uploaded_files.append(filepath)
                app.logger.info(f"  ✅ Загружен файл: {file.filename}")
        
        if not uploaded_files:
            return jsonify({'success': False, 'message': 'Не удалось загрузить файлы. Проверьте формат (должен быть .xlsx или .xls)'})
        
        # Читаем все группы из всех файлов
        all_groups = []
        seen_combinations = set()  # Для удаления дубликатов по (ID, username)
        seen_ids = set()  # Для проверки дубликатов по ID
        seen_usernames = set()  # Для проверки дубликатов по username
        
        total_before = 0
        duplicates_count = 0
        
        for file_path in uploaded_files:
            try:
                groups = TelegramSearcher.read_groups_from_excel(str(file_path))
                total_before += len(groups)
                
                for group in groups:
                    group_id = group.get('id')
                    username = group.get('username')
                    username_normalized = username.lower().strip() if username else None
                    
                    # Проверяем дубликаты: группа считается дубликатом, если совпадает ID ИЛИ username
                    duplicate_found = False
                    
                    # Проверка по ID
                    if group_id:
                        if group_id in seen_ids:
                            duplicate_found = True
                    
                    # Проверка по username (только если не найден дубликат по ID)
                    if not duplicate_found and username_normalized:
                        if username_normalized in seen_usernames:
                            duplicate_found = True
                    
                    # Если не дубликат - добавляем в результат и отмечаем как просмотренные
                    if not duplicate_found:
                        all_groups.append(group)
                        
                        # Добавляем в множества для проверки следующих групп
                        if group_id:
                            seen_ids.add(group_id)
                        if username_normalized:
                            seen_usernames.add(username_normalized)
                        if group_id and username_normalized:
                            seen_combinations.add((group_id, username_normalized))
                    else:
                        duplicates_count += 1
                        app.logger.debug(f"  🔄 Пропущен дубликат: ID={group_id}, username={username}")
                
                app.logger.info(f"  ✅ Из {file_path.name}: {len(groups)} групп ({len([g for g in groups if (g.get('id') not in seen_ids or (g.get('username') and g.get('username').lower() not in seen_usernames))])} уникальных)")
            except Exception as e:
                app.logger.error(f"  ❌ Ошибка чтения {file_path.name}: {e}")
        
        # Удаляем временные файлы
        for file_path in uploaded_files:
            try:
                if file_path.exists():
                    file_path.unlink()
            except Exception as e:
                app.logger.warning(f"Не удалось удалить временный файл {file_path}: {e}")
        
        if not all_groups:
            return jsonify({'success': False, 'message': 'Не удалось прочитать группы из файлов или все группы дубликаты'})
        
        # Сохраняем объединенный файл
        # Убеждаемся, что папка results существует
        results_dir = Path('results')
        results_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        result_filename = f'merged_groups_{timestamp}.xlsx'
        result_file = os.path.join('results', result_filename)
        
        # Определяем формат на основе первого файла (обычный или ready_format)
        # Проверяем наличие полей ready_format
        has_ready_format = any('check_status' in group or 'check_message' in group for group in all_groups[:5])
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Groups"
        
        if has_ready_format:
            headers = ['ID', 'Название', 'Username', 'Участников', 'Статус', 'Сообщение', 'Действие', 'Ключевое слово', 'Родительская группа']
        else:
            headers = ['ID', 'Название', 'Username', 'Количество участников', 'Ключевое слово']
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for group in all_groups:
            if has_ready_format:
                ws.append([
                    group.get('id', 'N/A'),
                    group.get('title', 'N/A'),
                    group.get('username') or 'N/A',
                    group.get('members_count', 'N/A'),
                    group.get('check_status', 'N/A'),
                    group.get('check_message', '') or 'N/A',
                    group.get('check_action', '') or 'N/A',
                    group.get('keyword', ''),
                    group.get('parent_group', 'N/A')
                ])
            else:
                ws.append([
                    group.get('id', 'N/A'),
                    group.get('title', 'N/A'),
                    group.get('username') or 'N/A',
                    group.get('members_count', 'N/A'),
                    group.get('keyword', '')
                ])
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(result_file)
        
        # Проверяем, что файл действительно создан
        if not os.path.exists(result_file):
            app.logger.error(f"❌ Файл не был создан: {result_file}")
            return jsonify({'success': False, 'message': 'Ошибка при сохранении файла'})
        
        app.logger.info(f"✅ Объединено {len(uploaded_files)} файлов: было {total_before} групп, стало {len(all_groups)} уникальных (удалено {duplicates_count} дубликатов)")
        app.logger.info(f"📁 Файл сохранен: {result_file} (размер: {os.path.getsize(result_file)} байт)")
        
        return jsonify({
            'success': True,
            'result_file': result_filename,
            'files_count': len(uploaded_files),
            'total_before': total_before,
            'total_after': len(all_groups),
            'duplicates_removed': duplicates_count
        })
        
    except Exception as e:
        app.logger.error(f"❌ Ошибка при объединении файлов: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

# Глобальные переменные для рассылки
sending_tasks = {}  # session_id -> task info
sending_stop_flags = {}  # session_id -> threading.Event

@app.route('/api/start_sending', methods=['POST'])
def start_sending():
    """Запуск рассылки сообщений"""
    session_id = get_session_id()
    
    try:
        # Получаем данные из формы
        filename = request.form.get('filename')  # Файл из списка
        uploaded_file = request.files.get('uploaded_file')  # Загруженный пользователем файл
        groups_text = request.form.get('groups_text', '').strip()  # Текстовый список аккаунтов
        message_text = request.form.get('message_text', '')
        message_limit = int(request.form.get('message_limit', 50))
        send_delay = float(request.form.get('send_delay', 5.0))
        
        # Получаем файлы
        photo_file = request.files.get('photo')
        video_file = request.files.get('video')
        
        # Определяем, какой источник использовать
        if uploaded_file:
            # Сохраняем загруженный файл
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            uploaded_filename = f'uploaded_groups_{timestamp}.xlsx'
            uploaded_filepath = os.path.join('results', uploaded_filename)
            os.makedirs('results', exist_ok=True)
            uploaded_file.save(uploaded_filepath)
            filename = uploaded_filename  # Используем загруженный файл
            app.logger.info(f"📁 Загружен файл пользователя: {uploaded_filename}")
        elif groups_text:
            # Создаем временный файл из текстового списка
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            temp_filename = f'text_groups_{timestamp}.xlsx'
            temp_filepath = os.path.join('results', temp_filename)
            os.makedirs('results', exist_ok=True)
            
            # Парсим текстовый список и создаем Excel файл
            groups_list = parse_groups_from_text(groups_text)
            if not groups_list:
                return jsonify({'success': False, 'message': 'Не удалось распознать аккаунты в тексте. Используйте формат: @username или username или ID группы, каждый с новой строки'})
            
            # Создаем Excel файл
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['ID', 'Название', 'Username', 'Участников', 'Ключевое слово'])
            
            for group in groups_list:
                group_id = group.get('id', '')
                username = group.get('username', '')
                title = group.get('title', username or f"ID: {group_id}" if group_id else 'N/A')
                
                ws.append([
                    group_id if group_id else '',  # ID (может быть пустым для username)
                    title,
                    username if username else '',  # Username (может быть пустым для ID)
                    '',
                    'text_input'
                ])
                
                app.logger.info(f"  📝 Добавлена группа: id={group_id}, username={username}, title={title}")
            
            wb.save(temp_filepath)
            filename = temp_filename
            app.logger.info(f"📝 Создан файл из текстового списка: {temp_filename} ({len(groups_list)} групп)")
        elif not filename:
            return jsonify({'success': False, 'message': 'Не выбран файл с группами, не загружен файл и не введен список аккаунтов'})
        
        if not message_text and not photo_file and not video_file:
            return jsonify({'success': False, 'message': 'Не указан текст сообщения или не загружены файлы'})
        
        # Загружаем API credentials из config.py
        spec = importlib.util.spec_from_file_location("config", "config.py")
        config = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(config)
        
        api_id = config.API_ID
        api_hash = config.API_HASH
        
        # Сохраняем загруженные файлы
        photo_path = None
        video_path = None
        
        if photo_file:
            photo_path = os.path.join('uploads', f'photo_{session_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg')
            os.makedirs('uploads', exist_ok=True)
            photo_file.save(photo_path)
        
        if video_file:
            video_path = os.path.join('uploads', f'video_{session_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.mp4')
            os.makedirs('uploads', exist_ok=True)
            video_file.save(video_path)
        
        # Создаем stop event
        stop_event = threading.Event()
        sending_stop_flags[session_id] = stop_event
        
        # Инициализируем задачу
        sending_tasks[session_id] = {
            'status': 'running',
            'progress': {
                'current': 0,
                'total': 0,
                'message': 'Инициализация...',
                'current_group': ''
            },
            'sent_count': 0,
            'error_count': 0,
            'blocked_count': 0,
            'skipped_count': 0,
            'logs': []
        }
        
        # Запускаем рассылку в отдельном потоке
        thread = threading.Thread(
            target=run_sending_async,
            args=(session_id, filename, message_text, message_limit, send_delay, photo_path, video_path, api_id, api_hash, stop_event)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'message': 'Рассылка запущена'})
        
    except Exception as e:
        app.logger.error(f"Ошибка при запуске рассылки: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'})

@app.route('/api/stop_sending', methods=['POST'])
def stop_sending():
    """Остановка рассылки"""
    session_id = get_session_id()
    stop_event = sending_stop_flags.get(session_id)
    
    if stop_event:
        stop_event.set()
        if session_id in sending_tasks:
            sending_tasks[session_id]['status'] = 'stopped'
        return jsonify({'success': True, 'message': 'Рассылка остановлена'})
    
    return jsonify({'success': False, 'message': 'Рассылка не найдена'})

@app.route('/api/sending_status', methods=['GET'])
def sending_status():
    """Получение статуса рассылки"""
    session_id = get_session_id()
    task = sending_tasks.get(session_id, {})
    progress = task.get('progress', {})
    
    response = {
        'status': task.get('status', 'idle'),
        'current': progress.get('current', 0),
        'total': progress.get('total', 0),
        'message': progress.get('message', ''),
        'current_group': progress.get('current_group', ''),
        'sent_count': task.get('sent_count', 0),
        'error_count': task.get('error_count', 0),
        'blocked_count': task.get('blocked_count', 0),
        'skipped_count': task.get('skipped_count', 0)
    }
    
    # Добавляем последний лог
    logs = task.get('logs', [])
    if logs:
        response['last_log'] = logs[-1]
    
    # Если завершено, добавляем файл отчета
    if task.get('status') == 'completed':
        response['report_file'] = task.get('report_file')
    
    return jsonify(response)

def run_sending_async(session_id, filename, message_text, message_limit, send_delay, photo_path, video_path, api_id, api_hash, stop_event):
    """Запуск рассылки в отдельном потоке"""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    
    try:
        loop.run_until_complete(
            send_messages_to_groups(session_id, filename, message_text, message_limit, send_delay, photo_path, video_path, api_id, api_hash, stop_event)
        )
    except Exception as e:
        app.logger.error(f"Ошибка в run_sending_async: {e}", exc_info=True)
        if session_id in sending_tasks:
            sending_tasks[session_id]['status'] = 'error'
            sending_tasks[session_id]['progress']['message'] = f'Ошибка: {str(e)}'
    finally:
        loop.close()

async def send_messages_to_groups(session_id, filename, message_text, message_limit, send_delay, photo_path, video_path, api_id, api_hash, stop_event):
    """Асинхронная функция рассылки сообщений"""
    try:
        # Используем уникальное имя сессии для каждой рассылки, чтобы избежать блокировки БД
        # Копируем основную сессию, если она существует, с повторными попытками
        main_session_path = 'telegram_session.session'
        session_name = f'telegram_session_{session_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        
        if os.path.exists(main_session_path):
            import shutil
            import time
            session_copy_path = f'{session_name}.session'
            
            # Пробуем скопировать сессию с повторными попытками
            for attempt in range(3):
                try:
                    # Небольшая задержка перед копированием
                    await asyncio.sleep(0.5)
                    shutil.copy2(main_session_path, session_copy_path)
                    app.logger.info(f"📋 Использую копию сессии: {session_name}")
                    break
                except Exception as e:
                    if attempt < 2:
                        app.logger.warning(f"⚠️ Попытка {attempt + 1} копирования сессии не удалась: {e}, повторяю...")
                        await asyncio.sleep(1)
                    else:
                        app.logger.warning(f"⚠️ Не удалось скопировать сессию после 3 попыток: {e}, использую новую сессию")
        
        searcher = TelegramSearcher(api_id, api_hash, session_name, send_delay)
        
        # Пробуем подключиться с таймаутом и повторными попытками
        connected = False
        for attempt in range(3):
            try:
                await asyncio.wait_for(searcher.client.start(), timeout=30.0)
                connected = True
                app.logger.info(f"✅ Успешно подключено к Telegram (попытка {attempt + 1})")
                break
            except asyncio.TimeoutError:
                if attempt < 2:
                    app.logger.warning(f"⏳ Таймаут подключения (попытка {attempt + 1}), повторяю...")
                    await asyncio.sleep(2)
                else:
                    sending_tasks[session_id]['status'] = 'error'
                    sending_tasks[session_id]['progress']['message'] = 'Таймаут подключения к Telegram после 3 попыток'
                    return
            except Exception as e:
                error_msg = str(e).lower()
                if 'database is locked' in error_msg and attempt < 2:
                    app.logger.warning(f"⏳ БД заблокирована (попытка {attempt + 1}), жду и повторяю...")
                    await asyncio.sleep(3)
                else:
                    app.logger.error(f"Ошибка подключения к Telegram: {e}")
                    sending_tasks[session_id]['status'] = 'error'
                    sending_tasks[session_id]['progress']['message'] = f'Ошибка подключения: {str(e)}'
                    return
        
        if not connected:
            sending_tasks[session_id]['status'] = 'error'
            sending_tasks[session_id]['progress']['message'] = 'Не удалось подключиться к Telegram'
            return
        
        # Читаем группы из файла
        file_path = os.path.join('results', filename)
        groups = TelegramSearcher.read_groups_from_excel(file_path)
        
        if not groups:
            sending_tasks[session_id]['status'] = 'error'
            sending_tasks[session_id]['progress']['message'] = 'Не удалось прочитать группы из файла. Проверьте формат файла.'
            app.logger.error(f"❌ Не удалось прочитать группы из файла: {file_path}")
            # Пробуем прочитать файл еще раз с детальным логированием
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                ws = wb.active
                app.logger.info(f"📄 Файл существует, строк: {ws.max_row}, колонок: {ws.max_column}")
                if ws.max_row > 1:
                    headers = [cell.value for cell in ws[1]]
                    app.logger.info(f"📋 Заголовки: {headers}")
                    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=min(5, ws.max_row), values_only=True)):
                        app.logger.info(f"  Строка {i+2}: {row}")
            except Exception as e:
                app.logger.error(f"❌ Ошибка при детальной проверке файла: {e}")
            return
        
        app.logger.info(f"📋 Прочитано {len(groups)} групп из файла: {filename}")
        
        # Логируем первую группу для отладки
        if groups:
            first_group = groups[0]
            app.logger.info(f"🔍 Первая группа: id={first_group.get('id')}, username={first_group.get('username')}, title={first_group.get('title')}")
        
        # Обновляем прогресс
        total_groups = min(len(groups), message_limit)
        sending_tasks[session_id]['progress'] = {
            'current': 0,
            'total': total_groups,
            'message': f'Начинаю рассылку в {total_groups} групп...',
            'current_group': ''
        }
        
        sent_count = 0
        error_count = 0
        blocked_count = 0
        skipped_count = 0
        results = []
        
        for i, group in enumerate(groups):
            if stop_event and stop_event.is_set():
                sending_tasks[session_id]['status'] = 'stopped'
                sending_tasks[session_id]['progress']['message'] = 'Рассылка остановлена пользователем'
                break
            
            # Проверяем лимит
            if sent_count + error_count >= message_limit:
                sending_tasks[session_id]['progress']['message'] = f'Достигнут лимит сообщений ({message_limit})'
                break
            
            group_title = group.get('title', f"ID: {group.get('id')}")
            group_id = group.get('id')
            username = group.get('username')
            
            # Логируем информацию о группе
            app.logger.info(f"📤 Обрабатываю группу: id={group_id}, username={username}, title={group_title}")
            
            # Обновляем прогресс
            sending_tasks[session_id]['progress'] = {
                'current': i + 1,
                'total': total_groups,
                'message': f'Отправляю в: {group_title}',
                'current_group': group_title
            }
            
            # Добавляем лог
            log_entry = {'message': f'Отправляю в: {group_title}', 'type': 'info'}
            sending_tasks[session_id]['logs'].append(log_entry)
            if len(sending_tasks[session_id]['logs']) > 100:
                sending_tasks[session_id]['logs'].pop(0)
            
            try:
                # Отправляем сообщение
                result = await searcher.send_message_to_group(
                    group_id, username, group_title, message_text, photo_path, video_path
                )
                
                if result['success']:
                    sent_count += 1
                    log_entry = {'message': f'✅ Отправлено в: {group_title}', 'type': 'success'}
                    results.append({
                        **group,
                        'status': 'sent',
                        'message': 'Сообщение отправлено успешно',
                        'timestamp': datetime.now().isoformat()
                    })
                elif result.get('blocked'):
                    blocked_count += 1
                    log_entry = {'message': f'🚫 Заблокировано: {group_title} - {result.get("message", "")}', 'type': 'error'}
                    results.append({
                        **group,
                        'status': 'blocked',
                        'message': result.get('message', 'Заблокировано'),
                        'timestamp': datetime.now().isoformat()
                    })
                else:
                    error_count += 1
                    log_entry = {'message': f'❌ Ошибка в {group_title}: {result.get("message", "")}', 'type': 'error'}
                    results.append({
                        **group,
                        'status': 'error',
                        'message': result.get('message', 'Ошибка отправки'),
                        'timestamp': datetime.now().isoformat()
                    })
                
                sending_tasks[session_id]['logs'].append(log_entry)
                if len(sending_tasks[session_id]['logs']) > 100:
                    sending_tasks[session_id]['logs'].pop(0)
                
            except Exception as e:
                error_count += 1
                error_msg = str(e)
                log_entry = {'message': f'❌ Исключение в {group_title}: {error_msg}', 'type': 'error'}
                sending_tasks[session_id]['logs'].append(log_entry)
                if len(sending_tasks[session_id]['logs']) > 100:
                    sending_tasks[session_id]['logs'].pop(0)
                
                results.append({
                    **group,
                    'status': 'error',
                    'message': error_msg,
                    'timestamp': datetime.now().isoformat()
                })
            
            # Обновляем счетчики
            sending_tasks[session_id]['sent_count'] = sent_count
            sending_tasks[session_id]['error_count'] = error_count
            sending_tasks[session_id]['blocked_count'] = blocked_count
            sending_tasks[session_id]['skipped_count'] = skipped_count
            
            # Задержка между отправками
            if i < len(groups) - 1:
                await asyncio.sleep(send_delay)
        
        # Сохраняем отчет
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f'sending_report_{timestamp}.xlsx'
        report_file = os.path.join('results', report_filename)
        
        searcher.save_sending_report(results, report_file, sent_count, error_count, blocked_count, skipped_count)
        
        # Завершаем
        sending_tasks[session_id]['status'] = 'completed'
        sending_tasks[session_id]['progress'] = {
            'current': total_groups,
            'total': total_groups,
            'message': 'Рассылка завершена',
            'current_group': ''
        }
        sending_tasks[session_id]['report_file'] = report_filename
        
        await searcher.disconnect()
        
        # Удаляем временную копию сессии
        session_copy_path = f'{session_name}.session'
        if os.path.exists(session_copy_path):
            try:
                os.remove(session_copy_path)
                app.logger.info(f"🗑️ Удалена временная сессия: {session_copy_path}")
            except Exception as e:
                app.logger.warning(f"⚠️ Не удалось удалить временную сессию: {e}")
        
        # Удаляем загруженные файлы
        if photo_path and os.path.exists(photo_path):
            os.remove(photo_path)
        if video_path and os.path.exists(video_path):
            os.remove(video_path)
        
    except Exception as e:
        app.logger.error(f"Ошибка в send_messages_to_groups: {e}", exc_info=True)
        if session_id in sending_tasks:
            sending_tasks[session_id]['status'] = 'error'
            sending_tasks[session_id]['progress']['message'] = f'Ошибка: {str(e)}'

if __name__ == '__main__':
    print("🚀 Запуск Flask приложения...")
    print("📱 Откройте в браузере: http://127.0.0.1:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)

